"""
Tests for ManualWorklistDESBackend (T2.9, spec.md section 7 items 1, 8, 9,
10, 11).

The formula tests reproduce the exact worked example verified against the
mentor's own ExpValues Eq sheet (section 7 item 8) using REAL training data
(row 0), not synthetic numbers. Two real corrections were caught this way
while building this file, not shipped silently:
  - an earlier version assumed the "Shift_..." columns held a 0-0.5
    fraction and multiplied by a baseline; the real data already holds the
    absolute shifted-volume value -- multiplying again would have been off
    by orders of magnitude.
  - a later version computed NA_Im_DR/etc via an addition formula
    (Shift_... + baseline); checked against real data, that value is
    already sitting in the training data as its own column (NA_Im_DR
    itself) -- the formula wasn't wrong, just unnecessary.
"""

from __future__ import annotations

import sys
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

import pandas as pd  # noqa: E402

from loop.des_backend.manual_worklist import (  # noqa: E402
    NO_ANYLOGIC_EQUIVALENT,
    ManualWorklistDESBackend,
    compute_raw_values,
    load_constants,
)


def _candidate(**overrides):
    base = {
        "NA_Im": 1.0, "NA_Ex": 1.0, "A_Im": 1.0, "A_Ex": 1.0,
        "Shift_NA_Im_LB_to_Cher": 0.0, "NA_Im_LB": 0.0, "NA_Im_DR": 0.0,
        "Shift_NA_Ex_LB_to_Cher": 0.0, "NA_Ex_LB": 0.0, "NA_Ex_DR": 0.0,
        "Shift_A_Im_LB_to_Cher": 0.0, "A_Im_LB": 0.0, "A_Im_DR": 0.0,
        "Shift_A_Ex_LB_to_Cher": 0.0, "A_Ex_LB": 0.0, "A_Ex_DR": 0.0,
        "VCap_Dub_Hey": 100.0, "VCap_Dub_Holy": 100.0, "VCap_Dub_Liv": 100.0,
        "VCap_Ross_Fish": 100.0, "VCap_Ross_Pem": 100.0,
        "ChkTime_Doc": 5.0, "ChkTime_Phy": 5.0,
        "NumCusShed_D": 3.0, "NumDAFM_D": 3.0, "NumCusShed_R": 3.0, "NumDAFM_R": 3.0,
        "Pct_NA_OB_Green": 0.7, "Pct_NA_OB_Red": 0.2, "Pct_A_OB_Red": 0.2,
        "Pct_NA_IB_Green": 0.7, "Pct_NA_IB_Red": 0.2, "Pct_A_IB_Red": 0.2,
        "Pct_IB_PreBoard": 0.2, "Pct_OB_PreBoard": 0.2,
    }
    base.update(overrides)
    return pd.Series(base)


class ComputeRawValuesFormulaTests(unittest.TestCase):
    def test_real_data_matches_mentor_sheet_exactly(self):
        """spec.md section 7 item 8's worked check, run through the real
        code path against the real first training row."""
        sys.path.insert(0, str(ROOT / "src"))
        sys.path.insert(0, str(ROOT.parent / "nolhc_ml" / "src"))
        from data import load_xy

        X_df, _ = load_xy()
        row = X_df.iloc[0]
        raw, _conflicts = compute_raw_values(row)

        # 6-Sep finding: these are non-percentage (Integer-typed in
        # AnyLogic) raw fields, now rounded to the nearest whole number
        # before being handed to a human to type in -- compare against the
        # ROUNDED expected value, not the pre-rounding mentor-sheet figure.
        self.assertEqual(raw["VolAgriImViaChe"], round(301281.22))
        self.assertEqual(raw["VolAgriImEULB"], round(127676.22))
        # current code path: direct passthrough of the NA_Im_DR column, not
        # a re-derivation -- confirmed to equal the formula's result too
        # (same underlying quantity, sourced two different ways in the raw
        # file), modulo the same integer rounding.
        self.assertEqual(raw["VolAllPImViaChe"], round(row["NA_Im_DR"]))
        self.assertEqual(raw["VolAllPImViaChe"], round(row["Shift_NA_Im_LB_to_Cher"] + 223000))

    def test_direct_passthrough(self):
        c = _candidate(NA_Im=42.0)
        raw, _ = compute_raw_values(c)
        self.assertEqual(raw["VolAllPImGB"], 42.0)

    def test_one_factor_fans_out_to_several_raw_names(self):
        c = _candidate(ChkTime_Doc=8.0)  # non-percentage factor -- use a value that survives rounding exactly
        raw, _ = compute_raw_values(c)
        for name in ["DocChkTimeAPImIR", "DocChkTimeAgriImIR", "DocChkTimeAPImGB-W",
                      "DocChkTimeAgriImGB-W", "DocCheckTimeImGB-E", "DocCheckTimeImEU"]:
            self.assertEqual(raw[name], 8)

    def test_non_percentage_raw_fields_round_to_nearest_integer(self):
        """6-Sep finding: manually entering an exported candidate's values
        into AnyLogic Cloud threw "must be integer" errors -- AnyLogic's
        own field types only allow decimals for the percentage/fraction
        fields. Every volume/count/capacity/check-time field must be a
        whole number."""
        c = _candidate(ChkTime_Doc=7.6, NumCusShed_D=3.4, VCap_Dub_Hey=100.5)
        raw, _ = compute_raw_values(c)
        self.assertEqual(raw["DocChkTimeAPImIR"], 8)  # 7.6 -> 8
        self.assertEqual(raw["NumCustomOfficerD"], 3)  # 3.4 -> 3
        self.assertIsInstance(raw["NumCustomOfficerD"], int)
        # percentage fields must NOT be rounded
        c2 = _candidate(Pct_NA_IB_Red=0.234, Pct_A_IB_Red=0.234)
        raw2, _ = compute_raw_values(c2)
        self.assertAlmostEqual(raw2["PerPhyChkImGB-E"], 0.234)

    def test_agreeing_shared_raw_names_produce_no_conflict(self):
        c = _candidate(Pct_NA_IB_Red=0.2, Pct_A_IB_Red=0.2)
        raw, conflicts = compute_raw_values(c)
        self.assertEqual(raw["PerPhyChkImGB-E"], 0.2)
        self.assertEqual(conflicts, {})

    def test_disagreeing_shared_raw_names_resolve_to_mean_and_are_reported(self):
        """Real finding (28-Aug, caught via real data): PerPhyChkImGB-E/EU
        are driven by BOTH Pct_NA_IB_Red and Pct_A_IB_Red, and real
        candidates routinely disagree (design point 1: 0.33 vs 0.28). This
        can't raise on every real candidate -- resolved to the mean, but
        the disagreement must show up in the conflicts report, not vanish
        silently the way a naive raw[name]=value overwrite would have done."""
        c = _candidate(Pct_NA_IB_Red=0.2, Pct_A_IB_Red=0.4)
        raw, conflicts = compute_raw_values(c)
        self.assertAlmostEqual(raw["PerPhyChkImGB-E"], 0.3)
        self.assertIn("PerPhyChkImGB-E", conflicts)
        contributing_factors = {factor for factor, _ in conflicts["PerPhyChkImGB-E"]}
        self.assertEqual(contributing_factors, {"Pct_NA_IB_Red", "Pct_A_IB_Red"})

    def test_no_anylogic_equivalent_factors_produce_no_raw_entry(self):
        c = _candidate(Pct_NA_OB_Green=0.9)
        raw, _ = compute_raw_values(c)
        self.assertNotIn("Pct_NA_OB_Green", raw)
        for name in raw:
            self.assertNotIn(name, NO_ANYLOGIC_EQUIVALENT)

    def test_three_factors_have_no_anylogic_equivalent(self):
        """Only the 3 outbound-percentage factors (spec.md section 7 item 9).
        Was 7 (4 more -- the landbridge-side volumes) from 28-Aug until
        6-Sep, when an independent live-dashboard check (a separate
        browser-automation tool, AUTOMATION_REPORT.md) showed those 4 DO
        map to real fields (the ViaRott columns) -- corrected, spec.md
        section 7 item 23. See FACTOR_TO_RAW for the real mapping."""
        self.assertEqual(
            NO_ANYLOGIC_EQUIVALENT,
            frozenset({"Pct_NA_OB_Green", "Pct_NA_OB_Red", "Pct_A_OB_Red"}),
        )

    def test_landbridge_volumes_map_to_the_via_rott_fields(self):
        """The 6-Sep correction itself: NA_Im_LB/NA_Ex_LB/A_Im_LB/A_Ex_LB
        now map to the confirmed real dashboard fields, not "no field"."""
        c = _candidate(NA_Im_LB=111.0, NA_Ex_LB=222.0, A_Im_LB=333.0, A_Ex_LB=444.0)
        raw, _ = compute_raw_values(c)
        self.assertEqual(raw["VolAllPImViaRott"], 111)
        self.assertEqual(raw["VolAllPExViaRott"], 222)
        self.assertEqual(raw["VolAgriImViaRott"], 333)
        self.assertEqual(raw["VolAgriExViaRott"], 444)


class ConstantsTests(unittest.TestCase):
    def test_loads_89_constants(self):
        constants = load_constants()
        self.assertEqual(len(constants), 89)
        self.assertIn("VolCatImGB", constants)
        self.assertEqual(constants["PerPhyChkLB"], 0)
        self.assertEqual(constants["PerSecurityChkLB"], 0)

    def test_missing_file_raises(self):
        with self.assertRaises(FileNotFoundError):
            load_constants(Path("/nonexistent/path.json"))


class ExportWorklistTests(unittest.TestCase):
    def test_export_creates_one_sheet_per_candidate_plus_cover(self):
        import openpyxl

        candidates = pd.DataFrame(
            [_candidate(NA_Im=10.0), _candidate(NA_Im=20.0)],
            index=["cand_a", "cand_b"],
        )
        backend = ManualWorklistDESBackend()
        out_dir = ROOT / "tests" / "_scratch"
        out_path = out_dir / "worklist_test.xlsx"
        backend.export_worklist(candidates, out_path)

        wb = openpyxl.load_workbook(out_path)
        self.assertEqual(set(wb.sheetnames), {"Cover", "cand_a", "cand_b"})
        ws = wb["cand_a"]
        header = [ws.cell(row=1, column=c).value for c in range(1, 4)]
        self.assertEqual(header, ["Field Name", "Value To Enter", "Type"])
        out_path.unlink()
        out_dir.rmdir()

    def test_export_empty_candidates_raises(self):
        backend = ManualWorklistDESBackend()
        with self.assertRaises(ValueError):
            backend.export_worklist(pd.DataFrame(), Path("/tmp/x.xlsx"))


class ExportRunRequestsCsvTests(unittest.TestCase):
    """The mentor-specified flat CSV (spec.md section 7 item 11): one row
    per run, 35 parameter columns (AnyLogic-named where possible), plus
    n_replications and seed."""

    def setUp(self):
        self.out_dir = ROOT / "tests" / "_scratch"
        self.out_path = self.out_dir / "requests_test.csv"

    def tearDown(self):
        if self.out_path.exists():
            self.out_path.unlink()
        if self.out_dir.exists():
            self.out_dir.rmdir()

    def test_one_row_per_candidate_with_run_id_and_35_columns_plus_meta(self):
        candidates = pd.DataFrame(
            [_candidate(NA_Im=10.0), _candidate(NA_Im=20.0)],
            index=["run_001", "run_002"],
        )
        backend = ManualWorklistDESBackend(constants={})
        out = backend.export_run_requests_csv(candidates, n_replications=5, seed=42, out_path=self.out_path)

        df = pd.read_csv(out)
        self.assertEqual(len(df), 2)
        self.assertEqual(list(df["run_id"]), ["run_001", "run_002"])
        self.assertEqual(list(df["n_replications"]), [5, 5])
        self.assertEqual(list(df["seed"]), [42, 42])
        # run_id + 35 parameter columns + n_replications + seed
        self.assertEqual(len(df.columns), 1 + 35 + 2)

    def test_column_headers_use_anylogic_names_where_known(self):
        candidates = pd.DataFrame([_candidate(NA_Im=99.0)], index=["run_a"])
        backend = ManualWorklistDESBackend(constants={})
        out = backend.export_run_requests_csv(candidates, n_replications=1, seed=None, out_path=self.out_path)

        df = pd.read_csv(out)
        self.assertIn("VolAllPImGB", df.columns)  # canonical AnyLogic name for NA_Im
        self.assertEqual(df.loc[0, "VolAllPImGB"], 99.0)
        # a no-AnyLogic-field factor keeps our own name
        self.assertIn("Pct_NA_OB_Green", df.columns)

    def test_missing_factor_column_raises(self):
        incomplete = pd.DataFrame([{"NA_Im": 1.0}], index=["run_a"])
        backend = ManualWorklistDESBackend(constants={})
        with self.assertRaises(ValueError):
            backend.export_run_requests_csv(incomplete, n_replications=1, seed=1, out_path=self.out_path)

    def test_empty_candidates_raises(self):
        backend = ManualWorklistDESBackend(constants={})
        with self.assertRaises(ValueError):
            backend.export_run_requests_csv(pd.DataFrame(), n_replications=1, seed=1, out_path=self.out_path)

    def test_invalid_n_replications_raises(self):
        candidates = pd.DataFrame([_candidate()], index=["run_a"])
        backend = ManualWorklistDESBackend(constants={})
        with self.assertRaises(ValueError):
            backend.export_run_requests_csv(candidates, n_replications=0, seed=1, out_path=self.out_path)

    def test_non_percentage_columns_are_integers_percentage_columns_stay_decimal(self):
        """6-Sep finding: a human hit "must be integer" errors in AnyLogic
        Cloud entering a previously-exported candidate's raw float values.
        Every non-percentage factor's CSV value must now be a whole number;
        the 8 percentage factors must NOT be rounded."""
        candidates = pd.DataFrame(
            [_candidate(NA_Im=5921300.242459027, Pct_NA_IB_Red=0.37741050016596667)],
            index=["run_a"],
        )
        backend = ManualWorklistDESBackend(constants={})
        out = backend.export_run_requests_csv(candidates, n_replications=1, seed=1, out_path=self.out_path)

        df = pd.read_csv(out)
        self.assertEqual(df.loc[0, "VolAllPImGB"], 5921300)  # NA_Im -- non-percentage, rounded
        self.assertAlmostEqual(df.loc[0, "PerPhyChkAPImIR"], 0.37741050016596667)  # Pct_NA_IB_Red's canonical column -- percentage, untouched


class IngestResultsTests(unittest.TestCase):
    def test_ingest_csv_round_trip(self):
        import tempfile

        df = pd.DataFrame({
            "run_id": ["cand_a", "cand_a"],
            "replication": [1, 2],
            "seed": [111, 112],
            "TT_OB_Agri": [30.1, 31.4],
            "Uti_DAFM_R": [0.15, 0.16],
        })
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
            df.to_csv(f.name, index=False)
            backend = ManualWorklistDESBackend(constants={})
            out = backend.ingest_results(f.name)

        self.assertEqual(list(out.columns), ["run_id", "replication", "seed", "tt_ob_agri", "uti_dafm_r"])
        self.assertEqual(len(out), 2)

    def test_missing_required_columns_raises(self):
        import tempfile

        df = pd.DataFrame({"TT_OB_Agri": [30.1]})
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
            df.to_csv(f.name, index=False)
            backend = ManualWorklistDESBackend(constants={})
            with self.assertRaises(ValueError):
                backend.ingest_results(f.name)

    def test_missing_file_raises(self):
        backend = ManualWorklistDESBackend(constants={})
        with self.assertRaises(FileNotFoundError):
            backend.ingest_results("/nonexistent/results.csv")


if __name__ == "__main__":
    unittest.main()
