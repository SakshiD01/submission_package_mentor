"""
One workbook: one sheet per target with SHAP metadata, mean |SHAP| table,
and embedded plots from ``run_step4_shap.py`` outputs.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import List, Set

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from step2_master_excel import _excel_sheet_title, _scaled_xl_image


def _safe_filename(s: str) -> str:
    return re.sub(r"[^a-zA-Z0-9_.-]+", "_", s)


def _unique_sheet_title(name: str, used: Set[str]) -> str:
    base = _excel_sheet_title(name)
    t = base
    n = 2
    while t in used:
        suf = f"_{n}"
        t = base[: max(1, 31 - len(suf))] + suf
        n += 1
    used.add(t)
    return t


def _shap_file_base(target: str, model: str) -> str:
    return f"{_safe_filename(target)}__{_safe_filename(model)}"


def build_shap_master_workbook(
    out_path: Path,
    *,
    shap_dir: Path,
    selected_csv: Path,
    max_width_px: int = 720,
) -> None:
    """
    Read ``shap_selected_models.csv`` and assemble ``*__importance.csv`` + PNGs
    from ``shap_dir`` into one ``.xlsx`` (one worksheet per target, same order as CSV).
    """
    if not selected_csv.is_file():
        raise FileNotFoundError(selected_csv)
    sel = pd.read_csv(selected_csv)
    need = {"target", "model", "selection_policy", "explain_split"}
    if not need.issubset(sel.columns):
        raise ValueError(f"{selected_csv} must have columns: {sorted(need)}")

    wb = Workbook()
    wb.remove(wb.active)
    titles_used: Set[str] = set()

    for _, row in sel.iterrows():
        target = str(row["target"])
        model = str(row["model"])
        base = _shap_file_base(target, model)
        stitle = _unique_sheet_title(target, titles_used)
        ws = wb.create_sheet(title=stitle)
        r = 1
        ws.cell(r, 1, f"SHAP — {target}").font = Font(bold=True, size=14)
        r += 2
        ws.cell(r, 1, "Model").font = Font(bold=True)
        ws.cell(r, 2, model)
        r += 1
        ws.cell(r, 1, "Selection policy").font = Font(bold=True)
        ws.cell(r, 2, str(row["selection_policy"]))
        r += 1
        ws.cell(r, 1, "Explain split").font = Font(bold=True)
        ws.cell(r, 2, str(row["explain_split"]))
        r += 2

        imp_path = shap_dir / f"{base}__importance.csv"
        if imp_path.is_file():
            ws.cell(r, 1, "Mean |SHAP| (explain subsample)").font = Font(bold=True)
            r += 1
            imp = pd.read_csv(imp_path)
            cols = list(imp.columns)
            for c, h in enumerate(cols, start=1):
                ws.cell(r, c, h).font = Font(bold=True)
            r += 1
            for _, ir in imp.iterrows():
                for c, h in enumerate(cols, start=1):
                    v = ir[h]
                    if pd.isna(v):
                        ws.cell(r, c, "")
                    elif isinstance(v, (int, float)) and not isinstance(v, bool):
                        ws.cell(r, c, float(v))
                    else:
                        ws.cell(r, c, str(v))
                r += 1
        else:
            ws.cell(r, 1, f"(Missing {imp_path.name} — SHAP may have failed for this target.)")
            r += 1

        r += 2
        ws.cell(r, 1, "Figures (embedded)").font = Font(bold=True)
        r += 2

        plot_paths: List[Path] = []
        for suffix in ("summary_bar", "beeswarm"):
            p = shap_dir / f"{base}__{suffix}.png"
            if p.is_file():
                plot_paths.append(p)
        wf = sorted(shap_dir.glob(f"{base}__waterfall_row*.png"))
        plot_paths.extend(wf)

        if not plot_paths:
            ws.cell(r, 1, f"(No PNGs found for prefix {base}__)")
            r += 1
        else:
            for p in plot_paths:
                xl_img = _scaled_xl_image(p, max_width_px=max_width_px)
                if xl_img:
                    ws.cell(r, 1, p.name).font = Font(italic=True, size=9)
                    r += 1
                    ws.add_image(xl_img, f"A{r}")
                    row_advance = max(28, int(xl_img.height / 18) + 2)
                    r += row_advance
                else:
                    ws.cell(r, 1, f"(Could not load {p})")
                    r += 2

        for col in range(1, 12):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = min(18 if col <= 2 else 14, 48)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
