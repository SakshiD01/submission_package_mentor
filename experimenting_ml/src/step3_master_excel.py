"""
One workbook: one sheet per target with Step 3 tables + embedded calibration / HP PNGs.
"""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

try:
    from PIL import Image as PILImage
except ImportError:
    PILImage = None  # type: ignore


def _excel_sheet_title(name: str) -> str:
    invalid = r"[]:*?/\\"
    s = "".join("_" if c in invalid else c for c in name)
    return s[:31] if len(s) > 31 else s


def _scaled_xl_image(path: Path, max_width_px: int) -> Optional[XLImage]:
    if not path.is_file():
        return None
    if PILImage is None:
        xl = XLImage(str(path))
        xl.width = max_width_px
        xl.height = int(xl.height * (max_width_px / max(xl.width, 1)))
        return xl
    pil = PILImage.open(path)
    if pil.mode == "RGBA":
        bg = PILImage.new("RGB", pil.size, (255, 255, 255))
        bg.paste(pil, mask=pil.split()[3])
        pil = bg
    elif pil.mode != "RGB":
        pil = pil.convert("RGB")
    w, h = pil.size
    if w > max_width_px:
        scale = max_width_px / w
        w = max_width_px
        h = int(h * scale)
        pil = pil.resize((w, h), PILImage.Resampling.LANCZOS)
    bio = BytesIO()
    pil.save(bio, format="PNG")
    bio.seek(0)
    xl = XLImage(bio)
    xl.width = w
    xl.height = h
    return xl


def _write_df(ws, df: pd.DataFrame, start_row: int, start_col: int = 1) -> int:
    if df is None or df.empty:
        ws.cell(start_row, start_col, "(no rows)")
        return start_row + 1
    for j, col in enumerate(df.columns, start=start_col):
        ws.cell(start_row, j, str(col)).font = Font(bold=True)
    r = start_row + 1
    for _, row in df.iterrows():
        for j, col in enumerate(df.columns, start=start_col):
            v = row[col]
            if pd.isna(v):
                ws.cell(r, j, "")
            elif isinstance(v, bool):
                ws.cell(r, j, v)
            else:
                ws.cell(r, j, v)
        r += 1
    return r


def build_step3_workbook(out_path: Path, step3_dir: Path) -> None:
    """
    Expects under ``step3_dir``:
      - per_target_cv_selection.csv
      - calibration_cv_metrics.csv (optional)
      - hp_sensitivity_run_log.csv (optional)
      - calibration/calibration__<target>__<model>.png
      - hp_sensitivity/hp_sens__<target>__<model>.png
    """
    sel_path = step3_dir / "per_target_cv_selection.csv"
    if not sel_path.is_file():
        raise FileNotFoundError(
            f"Missing {sel_path}; run run_step3_pre_conformal.py first."
        )

    sel = pd.read_csv(sel_path)
    targets = sorted(sel["target"].astype(str).unique())

    cal_path = step3_dir / "calibration_cv_metrics.csv"
    cal = pd.read_csv(cal_path) if cal_path.is_file() else pd.DataFrame()

    hp_path = step3_dir / "hp_sensitivity_run_log.csv"
    hp = pd.read_csv(hp_path) if hp_path.is_file() else pd.DataFrame()

    cal_img_dir = step3_dir / "calibration"
    hp_img_dir = step3_dir / "hp_sensitivity"

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    gen = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    for t in targets:
        ws = wb.create_sheet(title=_excel_sheet_title(t))
        row = 1
        c1 = ws.cell(row, 1, "Step 3 pre-conformal report")
        c1.font = Font(bold=True, size=12)
        ws.cell(row, 2, t)
        row += 1
        ws.cell(row, 1, "Generated (UTC)").font = Font(bold=True)
        ws.cell(row, 2, gen)
        ws.cell(row, 3, str(step3_dir.resolve())).font = Font(size=9)
        row += 2

        ws.cell(row, 1, "1. Per-target CV selection").font = Font(bold=True, size=11)
        row += 1
        ssub = sel[sel["target"].astype(str) == t]
        row = _write_df(ws, ssub, row, 1)
        row += 2

        ws.cell(row, 1, "2. Calibration CV metrics (OOF, top models)").font = Font(
            bold=True, size=11
        )
        row += 1
        if len(cal) and "target" in cal.columns:
            csub = cal[cal["target"].astype(str) == t].copy()
        else:
            csub = pd.DataFrame()
        row = _write_df(ws, csub, row, 1)
        row += 2

        ws.cell(row, 1, "3. HP sensitivity run log").font = Font(bold=True, size=11)
        row += 1
        if len(hp) and "target" in hp.columns:
            hsub = hp[hp["target"].astype(str) == t].copy()
        else:
            hsub = pd.DataFrame()
        row = _write_df(ws, hsub, row, 1)
        row += 2

        ws.cell(row, 1, "4. Calibration figures (top models)").font = Font(
            bold=True, size=11
        )
        row += 1
        if len(csub):
            models = csub["model"].astype(str).tolist()
            for m in models:
                p = cal_img_dir / f"calibration__{t}__{m}.png"
                xl = _scaled_xl_image(p, max_width_px=820)
                if xl:
                    ws.cell(row, 1, m).font = Font(italic=True, size=9)
                    row += 1
                    ws.add_image(xl, f"A{row}")
                    row += max(28, int(xl.height / 16) + 2)
                else:
                    ws.cell(row, 1, f"{m}: (PNG not found: {p.name})")
                    row += 2
        else:
            ws.cell(row, 1, "(Run Step 3 with calibration enabled; or no rows for this target.)")
            row += 2

        ws.cell(row, 1, "5. HP sensitivity figures").font = Font(bold=True, size=11)
        row += 1
        if len(hsub):
            for _, r in hsub.iterrows():
                m = str(r["model"])
                plotted = r.get("plotted", True)
                if isinstance(plotted, str):
                    plotted = plotted.lower() in ("true", "1", "yes")
                p = hp_img_dir / f"hp_sens__{t}__{m}.png"
                xl = _scaled_xl_image(p, max_width_px=820)
                if xl:
                    note = "plotted" if plotted else "plot may have failed (check log)"
                    ws.cell(row, 1, f"{m} ({note})").font = Font(italic=True, size=9)
                    row += 1
                    ws.add_image(xl, f"A{row}")
                    row += max(28, int(xl.height / 16) + 2)
                else:
                    ws.cell(
                        row,
                        1,
                        f"{m}: (PNG not found — need cv_fold_details + successful HP sweep)",
                    )
                    row += 2
        else:
            ws.cell(
                row,
                1,
                "(No HP log; run Step 1 with --save-fold-details and Step 3 without --skip-hp.)",
            )
            row += 2

        for c in range(1, 18):
            ws.column_dimensions[get_column_letter(c)].width = min(16, 22)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
