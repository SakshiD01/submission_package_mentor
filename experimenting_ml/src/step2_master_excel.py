"""
One workbook per run: one sheet per target with CV, Friedman/Nemenyi tables,
embedded CD diagram and learning-curve grid PNGs from Step 2 outputs.
"""

from __future__ import annotations

import json
import math
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

try:
    from PIL import Image as PILImage
except ImportError:
    PILImage = None  # type: ignore


def _safe_mean_rank(ranks: Dict[str, Any], m: str):
    v = ranks.get(m)
    if v is None:
        v = ranks.get(str(m))
    if v is None:
        return ""
    try:
        fv = float(v)
        if math.isnan(fv):
            return ""
        return fv
    except (TypeError, ValueError):
        return ""


def _excel_sheet_title(name: str) -> str:
    invalid = r"[]:*?/\\"
    s = "".join("_" if c in invalid else c for c in name)
    return s[:31] if len(s) > 31 else s


def _scaled_xl_image(path: Path, max_width_px: int) -> Optional[XLImage]:
    if not path.is_file():
        return None
    if PILImage is None:
        xl = XLImage(str(path))
        xl.width = max_width_px
        xl.height = int(xl.height * (max_width_px / max(xl.width, 1)))
        return xl
    pil = PILImage.open(path)
    if pil.mode == "RGBA":
        bg = PILImage.new("RGB", pil.size, (255, 255, 255))
        bg.paste(pil, mask=pil.split()[3])
        pil = bg
    elif pil.mode != "RGB":
        pil = pil.convert("RGB")
    w, h = pil.size
    if w > max_width_px:
        scale = max_width_px / w
        w = max_width_px
        h = int(h * scale)
        pil = pil.resize((w, h), PILImage.Resampling.LANCZOS)
    bio = BytesIO()
    pil.save(bio, format="PNG")
    bio.seek(0)
    xl = XLImage(bio)
    xl.width = w
    xl.height = h
    return xl


def _write_kv(ws, start_row: int, pairs: List[tuple]) -> int:
    r = start_row
    for k, v in pairs:
        ws.cell(r, 1, k).font = Font(bold=True)
        c = ws.cell(r, 2, v)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    return r


def _load_test_lookup(path: Optional[Path]) -> Optional[pd.DataFrame]:
    if path is None or not path.is_file():
        return None
    df = pd.read_csv(path)
    need = {"target", "model", "rmse", "mae", "r2"}
    if not need.issubset(set(df.columns)):
        return None
    return df[list(need)]


def build_master_workbook(
    out_path: Path,
    *,
    cv_path: Path,
    cv_results: Dict[str, Any],
    targets: List[str],
    model_names: List[str],
    step2_dir: Path,
    summary_df: pd.DataFrame,
    nemenyi_by_target: Dict[str, pd.DataFrame],
    meta: Dict[str, Any],
    alpha: float,
    test_results_csv: Optional[Path] = None,
) -> None:
    """
    Creates ``out_path`` with one worksheet per target (sanitized name).
    """
    test_df = _load_test_lookup(test_results_csv)
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    gen_time = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    cd_dir = step2_dir / "cd_per_target"
    lc_dir = step2_dir / "learning_curves"
    res_dir = step2_dir / "residual_plots"

    for target in targets:
        title = _excel_sheet_title(target)
        ws = wb.create_sheet(title=title)

        row = _write_kv(
            ws,
            1,
            [
                ("Target", target),
                ("Generated", gen_time),
                ("cv_results.json", str(cv_path.resolve())),
                ("Step 2 output dir", str(step2_dir.resolve())),
                ("Nemenyi alpha", str(alpha)),
                ("Method", str(meta.get("method", ""))),
            ],
        )
        row += 1

        sub = summary_df[summary_df["target"] == target]
        if len(sub) == 0:
            ws.cell(row, 1, "No Friedman summary row for this target.")
            row += 2
        else:
            srow = sub.iloc[0]
            row = _write_kv(
                ws,
                row,
                [
                    ("Friedman statistic", float(srow["friedman_statistic"])),
                    ("Friedman p-value", float(srow["friedman_p_value"])),
                    ("n_blocks (CV scores)", int(srow["n_blocks"])),
                    ("n_models", int(srow["n_models"])),
                    ("Best mean-rank model", str(srow["best_mean_rank_model"])),
                    ("Best mean rank", float(srow["best_mean_rank"])),
                    (
                        "Tier-1 (not sig. vs best, Nemenyi)",
                        str(srow["tier1_not_sig_vs_best"]),
                    ),
                ],
            )
            row += 2

        ws.cell(row, 1, "Per-model CV (Step 1) + mean rank (Friedman) + holdout test (if CSV)").font = Font(
            bold=True
        )
        row += 1
        hdr = [
            "model",
            "mean_cv_rmse",
            "std_cv_rmse",
            "cv_n_scores",
            "mean_rank",
            "test_rmse",
            "test_mae",
            "test_r2",
            "best_params_json",
        ]
        for c, h in enumerate(hdr, start=1):
            ws.cell(row, c, h).font = Font(bold=True)
        row += 1

        ranks = (meta.get(target) or {}).get("mean_ranks") or {}
        rows_data = []
        for m in model_names:
            r = cv_results[target][m]
            te_rmse = te_mae = te_r2 = ""
            if test_df is not None:
                hit = test_df[(test_df["target"] == target) & (test_df["model"] == m)]
                if len(hit):
                    te_rmse = float(hit.iloc[0]["rmse"])
                    te_mae = float(hit.iloc[0]["mae"])
                    te_r2 = float(hit.iloc[0]["r2"])
            rows_data.append(
                {
                    "model": m,
                    "mean_cv_rmse": float(r["mean_rmse"]),
                    "std_cv_rmse": float(r["std_rmse"]),
                    "cv_n_scores": int(
                        r.get("cv_n_scores", len(r.get("fold_rmses", [])))
                    ),
                    "mean_rank": _safe_mean_rank(ranks, m),
                    "test_rmse": te_rmse,
                    "test_mae": te_mae,
                    "test_r2": te_r2,
                    "best_params_json": json.dumps(r["best_params"], sort_keys=True),
                }
            )
        df_cv = pd.DataFrame(rows_data).sort_values("mean_cv_rmse")
        for _, sr in df_cv.iterrows():
            for c, h in enumerate(hdr, start=1):
                cell = ws.cell(row, c, sr[h])
                if h == "best_params_json":
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

        row += 2
        ws.cell(row, 1, "Nemenyi post-hoc p-values (pairwise)").font = Font(bold=True)
        row += 1
        nm = nemenyi_by_target.get(target)
        if nm is not None and len(nm):
            cols = [""] + list(nm.columns)
            for c, name in enumerate(cols, start=1):
                ws.cell(row, c, str(name)).font = Font(bold=True, size=8)
            row += 1
            for idx_name in nm.index:
                ws.cell(row, 1, str(idx_name)).font = Font(size=8)
                for c, colname in enumerate(nm.columns, start=2):
                    v = nm.loc[idx_name, colname]
                    try:
                        fv = float(v)
                        ws.cell(row, c, fv).font = Font(size=8)
                    except (TypeError, ValueError):
                        ws.cell(row, c, str(v)).font = Font(size=8)
                row += 1
        else:
            ws.cell(row, 1, "(missing)")
            row += 1

        row += 2
        ws.cell(row, 1, "Figures (embedded)").font = Font(bold=True)
        row += 1
        ws.cell(
            row,
            1,
            "CD diagram (average ranks); learning-curve grid (all models); "
            "residual-by-fold grid (all models) when Step 2 residuals were run.",
        ).alignment = Alignment(wrap_text=True)
        row += 2

        img_row_cd = row
        cd_png = cd_dir / f"{target}.png"
        xl_cd = _scaled_xl_image(cd_png, max_width_px=700)
        if xl_cd:
            ws.add_image(xl_cd, f"A{img_row_cd}")
            row += int(xl_cd.height / 20) + 2
        else:
            ws.cell(img_row_cd, 1, f"(CD PNG not found: {cd_png})")
            row = img_row_cd + 2

        grid_png = lc_dir / f"grid__{target}.png"
        xl_g = _scaled_xl_image(grid_png, max_width_px=950)
        if xl_g:
            ws.add_image(xl_g, f"A{row}")
            row += int(xl_g.height / 18) + 2
        else:
            ws.cell(row, 1, f"(Learning-curve grid not found: {grid_png})")
            row += 2

        res_grid_png = res_dir / f"grid_residuals__{target}.png"
        res_legacy_png = res_dir / f"residuals_by_fold__{target}.png"
        xl_r = _scaled_xl_image(res_grid_png, max_width_px=950)
        if xl_r is None:
            xl_r = _scaled_xl_image(res_legacy_png, max_width_px=800)
        if xl_r:
            ws.cell(row, 1, "CV residuals by fold (all models)").font = Font(bold=True)
            row += 1
            ws.add_image(xl_r, f"A{row}")

        for col in range(1, 10):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = min(14 + (3 if col == 9 else 0), 48)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
