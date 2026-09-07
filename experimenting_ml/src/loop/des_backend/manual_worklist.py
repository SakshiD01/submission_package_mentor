"""
T2.9 (spec.md §7 items 1, 8, 9, 10, 11): ManualWorklistDESBackend -- the
real AnyLogic Cloud loop, not the synthetic stand-in.

File format (mentor-specified, 28-Aug -- spec.md §7 item 11): flat CSV in
both directions, not the Excel worksheet this class originally shipped
with (v1, same day). Corrected here:

    export_run_requests_csv(candidates, n_replications, seed) -> CSV:
        one row per requested run -- run_id, the 35 parameter values (one
        column per NOLHC factor, named after its AnyLogic field where one
        exists), n_replications, seed. This is the canonical machine
        record of what to run.

    export_worklist(candidates) -> Excel (kept as a SECOND, optional
        artifact per the user's own choice, 28-Aug): every raw AnyLogic
        field (including the 89 constants) laid out for someone to follow
        while actually sitting at AnyLogic Cloud. Not the canonical
        record -- a convenience view derived from the same mapping.

    ingest_results(results_path) -> replication-level DataFrame (run_id,
        replication, seed, <kpi columns>), read from the CSV/Excel the
        human brings back -- unchanged, already matched spec.md §7 item 2's
        contract when this class was first built.

Column-naming correction (28-Aug, found while building the CSV export):
NA_Im_DR/NA_Ex_DR/A_Im_DR/A_Ex_DR are NOT computed via the shift-fraction
formula this class originally used (SHIFT_FACTOR_TO_DR_RAW, additive) --
checked against real data and they're already sitting in our own 35-column
training data as plain columns holding the exact final value (row 0:
A_Im_DR = 301281.22, matching the formula's result exactly, but as a
pre-existing column, not something to derive). The formula wasn't wrong
numerically, just unnecessary -- simplified to a direct passthrough like
everything else in FACTOR_TO_RAW. NA_Im_LB/NA_Ex_LB/A_Im_LB/A_Ex_LB (their
4 landbridge-side siblings) have NO known raw AnyLogic field at all
(consistent with the earlier finding that landbridge volume has no
separate input -- AnyLogic computes it as a residual), joining the
3-KPI-percentage NO_ANYLOGIC_EQUIVALENT set from item 9 -- 7 factors total
now have no AnyLogic field, not 3.
"""

from __future__ import annotations

import csv
import json
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd

_HERE = Path(__file__).resolve().parent
_REPO_ROOT = _HERE.parents[3]
_NOLHC_SRC = _REPO_ROOT / "nolhc_ml" / "src"
if str(_NOLHC_SRC) not in sys.path:
    sys.path.insert(0, str(_NOLHC_SRC))

from training_columns import col_to_slug  # noqa: E402

CONSTANTS_PATH = _REPO_ROOT / "nolhc_ml" / "data" / "raw" / "anylogic_manual_constants.json"

# Direct-passthrough factors: candidate's value IS the raw AnyLogic value,
# just possibly duplicated across several raw field names (spec.md's
# "Confirmed Varying" evidence, AnyLogic_Constants_Worklist.xlsx, plus the
# NA_*_DR/A_*_DR correction above). The FIRST name in each list is the
# canonical CSV column header (user's choice, 28-Aug) -- the rest are the
# other AnyLogic fields the same value also applies to, used by
# export_worklist()'s fuller per-field view, not the flat CSV.
FACTOR_TO_RAW: Dict[str, List[str]] = {
    "NA_Im": ["VolAllPImGB"],
    "NA_Ex": ["VolAllPExGB"],
    "A_Im": ["VolAgriImGB"],
    "A_Ex": ["VolAgriExGB"],
    "Shift_NA_Im_LB_to_Cher": ["VolAllPImEULB"],
    "Shift_NA_Ex_LB_to_Cher": ["VolAllPExEULB"],
    "Shift_A_Im_LB_to_Cher": ["VolAgriImEULB"],
    "Shift_A_Ex_LB_to_Cher": ["VolAgriExEULB"],
    "NA_Im_DR": ["VolAllPImViaChe"],
    "NA_Ex_DR": ["VolAllPExViaChe"],
    "A_Im_DR": ["VolAgriImViaChe"],
    "A_Ex_DR": ["VolAgriExViaChe"],
    "VCap_Dub_Hey": ["DToHeyVesselCap"],
    "VCap_Dub_Holy": ["DToHolyVesselCap"],
    "VCap_Dub_Liv": ["DToLivVesselCap"],
    "VCap_Ross_Fish": ["RToFishVesselCap"],
    "VCap_Ross_Pem": ["RToPemVesselCap"],
    "ChkTime_Doc": ["DocChkTimeAPImIR", "DocChkTimeAgriImIR", "DocChkTimeAPImGB-W",
                     "DocChkTimeAgriImGB-W", "DocCheckTimeImGB-E", "DocCheckTimeImEU"],
    "ChkTime_Phy": ["PhyChkTimeAPImIR", "PhyChkTimeAgriImIR", "PhyChkTimeAPImGB-W",
                     "PhyChkTimeAgriImGB-W", "PhyCheckTimeImGB-E", "PhyCheckTimeImEU"],
    "NumCusShed_D": ["NumCustomOfficerD"],
    "NumDAFM_D": ["NumDAFMOfficerD"],
    "NumCusShed_R": ["NumCustomOfficerR"],
    "NumDAFM_R": ["NumDAFMOfficerR"],
    "Pct_NA_IB_Green": ["PerGreenTrucksAPImIR", "PerGreenTrucksAPImGB-W",
                         "PerGreenTrucksImGB-E", "PerGreenTrucksImEU"],
    "Pct_NA_IB_Red": ["PerPhyChkAPImIR", "PerPhyChkAPImGB-W", "PerPhyChkImGB-E", "PerPhyChkImEU"],
    "Pct_A_IB_Red": ["PerPhyChkAgriImIR", "PerPhyChkAgriImGB-W", "PerPhyChkImGB-E", "PerPhyChkImEU"],
    "Pct_IB_PreBoard": ["PerSecurityChkAPIR", "PerSecurityChkAPGB-W",
                         "PerSecurityChkGB-E", "PerSecurityChkEU"],
    "Pct_OB_PreBoard": ["PerSecurityChkAgriIR", "PerSecurityChkAgriGB-W",
                         "PerSecurityChkGB-E", "PerSecurityChkEU"],
    # Correction, 6-Sep (spec.md §7 item 23): these 4 were wrongly marked
    # NO_ANYLOGIC_EQUIVALENT below -- that was a reasoned inference from
    # 28-Aug ("landbridge volume must be a residual AnyLogic computes
    # internally"), never a live dashboard check. A separate, independently
    # -built browser-automation tool inspected the real, live dashboard
    # field list (discovered/inputs.json) and confirmed these ARE real
    # fields, under different names than our own -- the Rotterdam-direct
    # ("ViaRott") volumes, not a residual at all. Confirmed by more than a
    # label match: the tool successfully set these exact fields on 10 real
    # candidates and completed all 10 runs (round_20260906_090349,
    # AUTOMATION_REPORT.md §7.1/§9). Every worklist exported before this fix
    # told a human to skip these 4 fields -- real, retroactive concern for
    # any round entered by hand before today, spec.md §7 item 23.
    "NA_Im_LB": ["VolAllPImViaRott"],
    "NA_Ex_LB": ["VolAllPExViaRott"],
    "A_Im_LB": ["VolAgriImViaRott"],
    "A_Ex_LB": ["VolAgriExViaRott"],
}

# One canonical AnyLogic column name per factor, for the flat CSV (32
# factors -- the other 3 have no AnyLogic field at all, see below).
CANONICAL_CSV_COLUMN: Dict[str, str] = {factor: raw_names[0] for factor, raw_names in FACTOR_TO_RAW.items()}

# Confirmed: only 3 of our 35 factors have no AnyLogic field to write to --
# 3 outbound truck-check percentages (user-confirmed 28-Aug, spec.md §7 item 9).
# Was 7 (4 more, the landbridge-side volumes) until 6-Sep -- corrected,
# spec.md §7 item 23, once a live dashboard check showed those 4 DO map to
# real fields (the ViaRott columns), just not what this repo assumed.
# The flat CSV still includes a column for each (per "35 parameter values"),
# headed by OUR OWN factor name since there's no AnyLogic name to use --
# documented in spec.md as "present for completeness, leave blank / ignore
# in AnyLogic."
NO_ANYLOGIC_EQUIVALENT = frozenset({
    "Pct_NA_OB_Green", "Pct_NA_OB_Red", "Pct_A_OB_Red",
})

ALL_35_FACTORS: List[str] = sorted(set(FACTOR_TO_RAW) | NO_ANYLOGIC_EQUIVALENT)

# Real finding, 6-Sep: manually typing an exported candidate's values into
# AnyLogic Cloud threw "must be integer" errors. AnyLogic's own field types
# only allow decimals for the percentage/fraction fields -- every volume,
# staff count, vessel capacity, and check-duration field is Integer-typed,
# not Double. Our candidate proposer samples every factor as a continuous
# float (uniform-random within each factor's observed range), so every
# non-percentage factor needs rounding before it's written anywhere a human
# will copy it into AnyLogic -- it was never rounded before this.
#
# 8 of the 35 factors are the percentage/fraction ones (decimal-allowed):
# the 5 inbound-side ones in FACTOR_TO_RAW (their raw AnyLogic field names
# all start with "Per" -- PerPhyChk*/PerSecurityChk*/PerGreenTrucks*, a
# reliable, checked naming convention across every entry in that dict) plus
# the 3 outbound-side ones that have no AnyLogic field at all. Every other
# factor (volumes, vessel capacities, staff counts, check-time minutes) is
# Integer-typed in AnyLogic and gets rounded to the nearest whole number.
PERCENTAGE_FACTORS: frozenset = frozenset(
    {f for f, raw_names in FACTOR_TO_RAW.items() if raw_names[0].startswith("Per")}
    | {"Pct_NA_OB_Green", "Pct_NA_OB_Red", "Pct_A_OB_Red"}
)


def _anylogic_value(factor: str, value: float):
    """The value to actually write for one factor -- decimal for the 8
    percentage factors, rounded to a plain int for the other 27 (AnyLogic's
    own field-type constraint, not a choice we're making). Returns a real
    Python int, not a rounded float, so the CSV/worksheet shows "5921300"
    rather than "5921300.0" -- the latter still reads as ambiguous to
    whoever's typing it into an Integer-typed AnyLogic field by hand."""
    if factor in PERCENTAGE_FACTORS:
        return value
    return int(round(value))


def _is_percentage_raw_field(raw_field_name: str) -> bool:
    """Same rule as PERCENTAGE_FACTORS, applied to a raw AnyLogic field name
    directly (for compute_raw_values()'s output, which is keyed by raw field
    name, not our own factor name) -- every percentage raw field in
    FACTOR_TO_RAW starts with "Per", checked against the full dict above."""
    return raw_field_name.startswith("Per")


def load_constants(path: Optional[Path] = None) -> Dict[str, float]:
    p = Path(path) if path is not None else CONSTANTS_PATH
    if not p.is_file():
        raise FileNotFoundError(f"Missing constants file: {p}")
    with open(p, encoding="utf-8") as f:
        return json.load(f)["constants"]


def compute_raw_values(candidate: pd.Series) -> Tuple[Dict[str, float], Dict[str, List[Tuple[str, float]]]]:
    """One candidate's 35 NOLHC factor values -> (raw values, conflicts) for
    every raw AnyLogic field FACTOR_TO_RAW knows about (used by
    export_worklist()'s full per-field view -- the flat CSV only needs the
    35 canonical values, computed directly from the candidate row instead).

    Two of our factors legitimately share raw fields: Pct_NA_IB_Red and
    Pct_A_IB_Red both drive PerPhyChkImGB-E/PerPhyChkImEU (the user's own
    resolution table, spec.md §7 item 1 -- Dover/Calais checkpoints don't
    distinguish agri/non-agri, only Irish/west-GB ports do). Checked against
    real data (28-Aug): this is NOT a rare edge case -- design point 1 alone
    has Pct_NA_IB_Red=0.33 vs. Pct_A_IB_Red=0.28, and most real candidates
    will disagree the same way. Raising on every such candidate would make
    this unusable, so the resolution policy is: use the MEAN of the
    disagreeing values for the raw field, and report every conflict in the
    second return value so it's visible on the worklist, not silently
    averaged away (spec.md §7 item 10)."""
    contributions: Dict[str, List[Tuple[str, float]]] = {}
    for factor, raw_names in FACTOR_TO_RAW.items():
        if factor not in candidate.index:
            continue
        value = candidate[factor]
        for name in raw_names:
            contributions.setdefault(name, []).append((factor, value))

    raw: Dict[str, float] = {}
    conflicts: Dict[str, List[Tuple[str, float]]] = {}
    for name, contribs in contributions.items():
        values = {v for _, v in contribs}
        if len(values) > 1:
            conflicts[name] = contribs
            value = sum(v for _, v in contribs) / len(contribs)
        else:
            value = contribs[0][1]
        # AnyLogic's own field-type constraint (6-Sep finding, see
        # PERCENTAGE_FACTORS above) -- every non-percentage raw field is
        # Integer-typed, so round it here, after any conflict-averaging,
        # not before (averaging two already-rounded values could otherwise
        # produce a spurious .5).
        raw[name] = value if _is_percentage_raw_field(name) else int(round(value))

    return raw, conflicts


class ManualWorklistDESBackend:
    def __init__(self, constants: Optional[Dict[str, float]] = None) -> None:
        self.constants = constants if constants is not None else load_constants()

    def export_run_requests_csv(
        self,
        candidates: pd.DataFrame,
        n_replications: int,
        seed: Optional[int],
        out_path: Path,
    ) -> Path:
        """Mentor-specified flat CSV (spec.md §7 item 11): one row per
        requested run -- run_id, the 35 parameter values (AnyLogic column
        names where one exists, our own factor name otherwise), n_replications,
        seed. This is the canonical machine-readable request record.

        Values are rounded per PERCENTAGE_FACTORS (6-Sep finding): 27 of the
        35 factors are Integer-typed fields in AnyLogic Cloud and were
        previously written as raw floats, which AnyLogic's manual-entry form
        rejects outright ("must be integer"). Only the 8 percentage/fraction
        factors keep their decimal value."""
        if candidates.empty:
            raise ValueError("candidates is empty")
        if n_replications < 1:
            raise ValueError(f"n_replications must be >= 1, got {n_replications}")
        missing_factors = [f for f in ALL_35_FACTORS if f not in candidates.columns]
        if missing_factors:
            raise ValueError(f"candidates missing required factor columns: {missing_factors}")

        out_path = Path(out_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)

        header = ["run_id"] + [
            CANONICAL_CSV_COLUMN.get(f, f) for f in ALL_35_FACTORS
        ] + ["n_replications", "seed"]

        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(header)
            for run_id, row in candidates.iterrows():
                values = [_anylogic_value(f, row[f]) for f in ALL_35_FACTORS]
                writer.writerow([run_id] + values + [n_replications, seed])

        return out_path

    def import_run_requests_csv(self, path: Path) -> pd.DataFrame:
        """Inverse of export_run_requests_csv() -- recovers the 35-column,
        factor-named candidate DataFrame (indexed by run_id) from a
        previously-exported request CSV. Needed because a real manual round
        spans two separate process invocations (export today, ingest once
        the human brings back AnyLogic Cloud results days later) -- the
        in-memory flagged_batch from export_manual_round() doesn't survive
        that gap, so the CLI's ingest step reconstructs it from the file
        instead of requiring the caller to re-pickle a DataFrame by hand."""
        path = Path(path)
        if not path.is_file():
            raise FileNotFoundError(f"Missing run requests CSV: {path}")
        df = pd.read_csv(path)

        header_to_factor = {CANONICAL_CSV_COLUMN.get(f, f): f for f in ALL_35_FACTORS}
        missing = [h for h in header_to_factor if h not in df.columns]
        if missing:
            raise ValueError(f"run requests CSV missing expected columns: {missing}")
        if "run_id" not in df.columns:
            raise ValueError("run requests CSV missing 'run_id' column")

        factor_df = df.rename(columns=header_to_factor)[list(header_to_factor.values())]
        factor_df.index = df["run_id"]
        factor_df.index.name = None
        return factor_df

    def export_worklist(self, candidates: pd.DataFrame, out_path: Path) -> Path:
        """Second, optional artifact (kept per the user's choice, 28-Aug):
        one Excel sheet per candidate, every raw AnyLogic field to enter
        (including the 89 constants) and whether it varies or is constant --
        for whoever is actually sitting in front of AnyLogic Cloud. The
        canonical request record is export_run_requests_csv(); this is a
        derived convenience view, not a replacement for it."""
        if candidates.empty:
            raise ValueError("candidates is empty")

        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        varying_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        conflict_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

        cover = wb.create_sheet("Cover")
        cover["A1"] = "Manual AnyLogic Cloud Worklist"
        cover["A1"].font = Font(bold=True, size=14)
        cover["A2"] = f"{len(candidates)} candidate(s) -- one sheet each. Enter every row's value into the matching AnyLogic Cloud field, run, export results."
        na_factors = [f for f in candidates.columns if f in NO_ANYLOGIC_EQUIVALENT]
        cover["A4"] = "Do NOT enter anything for these factors -- no AnyLogic field exists (spec.md section 7 items 9 & 11):"
        row_cursor = 5
        for f in na_factors:
            cover.cell(row=row_cursor, column=1, value=f"  - {f}")
            row_cursor += 1

        row_cursor += 1
        cover.cell(row=row_cursor, column=1,
                    value="Value conflicts (2 NOLHC factors share 1 AnyLogic field, values differed -- resolved to their mean; see per-candidate sheets, red highlight):")
        row_cursor += 1
        any_conflicts = False
        for run_id, row in candidates.iterrows():
            _, conflicts = compute_raw_values(row)
            for name, contribs in conflicts.items():
                any_conflicts = True
                detail = ", ".join(f"{factor}={value}" for factor, value in contribs)
                cover.cell(row=row_cursor, column=1, value=f"  - {run_id}: {name} <- {detail}")
                row_cursor += 1
        if not any_conflicts:
            cover.cell(row=row_cursor, column=1, value="  (none)")
        cover.column_dimensions["A"].width = 100

        for run_id, row in candidates.iterrows():
            sheet_name = str(run_id)[:31] or f"candidate_{len(wb.sheetnames)}"
            ws = wb.create_sheet(sheet_name)
            headers = ["Field Name", "Value To Enter", "Type"]
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.fill = header_fill
                cell.font = header_font

            raw_varying, conflicts = compute_raw_values(row)
            r = 2
            for name, value in sorted(raw_varying.items()):
                ws.cell(row=r, column=1, value=name)
                ws.cell(row=r, column=2, value=value)
                if name in conflicts:
                    detail = ", ".join(f"{factor}={v}" for factor, v in conflicts[name])
                    cell = ws.cell(row=r, column=3, value=f"CONFLICT (mean of {detail})")
                    cell.fill = conflict_fill
                else:
                    cell = ws.cell(row=r, column=3, value="VARIES (this candidate)")
                    cell.fill = varying_fill
                r += 1
            for name, value in sorted(self.constants.items()):
                ws.cell(row=r, column=1, value=name)
                ws.cell(row=r, column=2, value=value)
                ws.cell(row=r, column=3, value="constant")
                r += 1

            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 18
            ws.column_dimensions["C"].width = 40
            ws.freeze_panes = "A2"

        out_path = Path(out_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.properties.creator = ""
        wb.properties.lastModifiedBy = ""
        wb.save(out_path)
        return out_path

    def ingest_results(self, results_path: Path) -> pd.DataFrame:
        """Reads back the human's AnyLogic Cloud export. Expected columns:
        run_id, replication, seed, then one column per KPI (raw_key names,
        e.g. 'TT_OB_Agri') -- matching spec.md §7 item 2's contract, the
        same shape SyntheticDESBackend.simulate() returns so downstream
        retrain code doesn't care which backend produced it."""
        results_path = Path(results_path)
        if not results_path.is_file():
            raise FileNotFoundError(f"Missing results file: {results_path}")

        if results_path.suffix.lower() == ".csv":
            df = pd.read_csv(results_path)
        else:
            df = pd.read_excel(results_path)

        required = {"run_id", "replication", "seed"}
        missing = required - set(df.columns)
        if missing:
            raise ValueError(f"results file missing required columns: {sorted(missing)}")
        if df.empty:
            raise ValueError("results file has no rows")

        kpi_cols = [c for c in df.columns if c not in required]
        rename = {c: col_to_slug(c) for c in kpi_cols}
        return df.rename(columns=rename)
