"""
One workbook: one sheet per target with Step 7 (test) + Step 9 (conformal) tables.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from step3_master_excel import _excel_sheet_title, _scaled_xl_image, _write_df

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
BEST_FILL = PatternFill("solid", fgColor="E2EFDA")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=12, color="1F4E79")


def _merged_table_for_target(
    target: str,
    test_results: Dict[str, Any],
    conformal_results: Dict[str, Any],
    model_names: List[str],
) -> pd.DataFrame:
    rows = []
    for m in model_names:
        if m not in test_results[target] or m not in conformal_results[target]:
            continue
        t = test_results[target][m]
        c = conformal_results[target][m]
        rows.append(
            {
                "model": m,
                "test_rmse": t["rmse"],
                "test_mae": t["mae"],
                "test_r2": t["r2"],
                "conformal_coverage_level": c["coverage_level"],
                "relative_rmse_to_best": c.get("relative_rmse_to_best", ""),
                "conformal_quantile_abs_res": c["quantile"],
                "empirical_interval_coverage": c["empirical_coverage"],
                "interval_width": c["interval_width"],
            }
        )
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    df = df.sort_values("test_rmse", ascending=True).reset_index(drop=True)
    df.insert(0, "rank_by_test_rmse", range(1, len(df) + 1))
    best = float(df["test_rmse"].min())
    df["is_best_test_rmse"] = df["test_rmse"].apply(lambda x: "Yes" if x == best else "")
    return df


def _adaptive_conformal_selected_df(
    out_dir: Path,
    test_results: Dict[str, Any],
    conformal_results: Dict[str, Any],
    targets: List[str],
    model_names: List[str],
) -> Tuple[Optional[pd.DataFrame], str]:
    """
    Prefer ``test_evaluation_final/selected_model_adaptive_conformal_from_step9.csv``;
    else build the same table from composite model selection when cv + paired tests exist.
    """
    csv_path = out_dir / "test_evaluation_final" / "selected_model_adaptive_conformal_from_step9.csv"
    if csv_path.is_file():
        return pd.read_csv(csv_path), f"loaded from {csv_path.name}"

    cv_path = out_dir / "cv_results.json"
    pairs_path = out_dir / "paired_ttests_all_targets.csv"
    if not cv_path.is_file() or not pairs_path.is_file():
        return None, (
            f"add {csv_path.name} (run run_test_set_evaluation_final.py) or ensure "
            f"{cv_path.name} and {pairs_path.name} for inline composite selection"
        )

    from excel_report import build_ttest_structures, compute_model_selection, load_json

    cv_results = load_json(cv_path)
    ttest_by_target = build_ttest_structures(pairs_path, model_names, targets)
    sel = compute_model_selection(
        cv_results,
        test_results,
        conformal_results,
        ttest_by_target,
        targets,
        model_names,
    )
    rows = []
    for _, r in sel.iterrows():
        t = str(r["Target"])
        m = str(r["Best_Model"])
        cr = conformal_results[t][m]
        rows.append(
            {
                "target": t,
                "selected_model": m,
                "adaptive_coverage_level": cr["coverage_level"],
                "empirical_coverage": cr["empirical_coverage"],
                "interval_width": cr["interval_width"],
                "relative_rmse_to_best": cr.get("relative_rmse_to_best", ""),
            }
        )
    return pd.DataFrame(rows), "computed from composite selection (matches CSV from test_evaluation_final)"


def build_step7_9_master_workbook(
    out_path: Path,
    *,
    test_results: Dict[str, Any],
    conformal_results: Dict[str, Any],
    targets: List[str],
    model_names: List[str],
    plots_dir: Optional[Path] = None,
    out_dir: Optional[Path] = None,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    wm = wb.create_sheet("Methodology", 0)
    wm.cell(1, 1, "Steps 7 & 9 — hold-out test metrics + adaptive conformal intervals").font = TITLE_FONT
    lines = [
        "",
        "Each target sheet: all models with test RMSE/MAE/R² and conformal summaries (spec §9).",
        "Adaptive coverage_level (0.90 / 0.95 / 0.99) depends on relative test RMSE vs best model per target.",
        "Symmetric intervals: ± quantile of |test residuals| at that coverage; empirical_coverage on the hold-out set.",
        "",
        "is_best_test_rmse = Yes for the lowest test_rmse in that target block.",
        "",
        "Sheet 'Adaptive_Conformal_Selected': Step 9 adaptive intervals for the composite-selected "
        "model per target (same as selected_model_adaptive_conformal_from_step9.csv when present).",
    ]
    for i, line in enumerate(lines, start=2):
        wm.cell(i, 1, line)
        wm.cell(i, 1).alignment = Alignment(wrap_text=True, vertical="top")
    wm.column_dimensions["A"].width = 88

    summary_rows = []

    for target in targets:
        if target not in test_results:
            continue
        df = _merged_table_for_target(
            target, test_results, conformal_results, model_names
        )
        st = _excel_sheet_title(target)
        ws = wb.create_sheet(title=st)
        ws.cell(1, 1, f"Target: {target}").font = TITLE_FONT
        n_test = 0
        if model_names and model_names[0] in test_results[target]:
            yp = test_results[target][model_names[0]].get("y_pred") or []
            n_test = len(yp)
        ws.cell(2, 1, f"Hold-out test rows (n): {n_test}")
        ws.cell(3, 1, "Sorted by test_rmse (ascending).")

        start = 5
        if df.empty:
            ws.cell(start, 1, "(no rows)")
        else:
            _write_df(ws, df, start_row=start, start_col=1)
            best_row = start + 1 + int(df["test_rmse"].idxmin())
            for ci in range(1, len(df.columns) + 1):
                ws.cell(best_row, ci).fill = BEST_FILL
            br = df.loc[df["test_rmse"].idxmin()]
            summary_rows.append(
                {
                    "target": target,
                    "best_model_test_rmse": br["model"],
                    "best_test_rmse": br["test_rmse"],
                    "best_test_r2": br["test_r2"],
                    "conformal_level_at_best": br["conformal_coverage_level"],
                    "interval_width_at_best": br["interval_width"],
                }
            )

        # Optional residual diagnostic figure from run_test_set_evaluation_final.py
        if plots_dir is not None:
            from re import sub

            safe = sub(r"[^a-zA-Z0-9_.-]+", "_", target)
            png = plots_dir / f"{safe}_holdout_residual_diagnostics.png"
            xl_img = _scaled_xl_image(png, max_width_px=900)
            if xl_img is not None:
                # Table: header at ``start``, data through ``start + len(df)``
                r_img = (start + len(df) + 2) if not df.empty else (start + 2)
                ws.cell(r_img, 1, "Residual diagnostics (from test_evaluation_final/plots):").font = Font(
                    bold=True
                )
                ws.add_image(xl_img, f"A{r_img + 1}")

        for col in range(1, 14):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = min(14, 22)

    ws_sum = wb.create_sheet("Summary", 1)
    ws_sum.cell(1, 1, "Best per target (lowest test RMSE)").font = TITLE_FONT
    if summary_rows:
        sdf = pd.DataFrame(summary_rows)
        _write_df(ws_sum, sdf, start_row=3, start_col=1)
    ws_sum.column_dimensions["A"].width = 22
    for col in range(2, 8):
        ws_sum.column_dimensions[get_column_letter(col)].width = 16

    # --- Adaptive conformal for composite-selected model only (CSV or inline) ---
    od = out_dir if out_dir is not None else out_path.parent
    adapt_df, adapt_note = _adaptive_conformal_selected_df(
        od, test_results, conformal_results, targets, model_names
    )
    ws_ad = wb.create_sheet("Adaptive_Conformal_Selected")
    ws_ad.cell(1, 1, "Step 9 adaptive conformal — composite-selected model per target").font = TITLE_FONT
    ws_ad.cell(2, 1, adapt_note).font = Font(italic=True, size=9)
    ws_ad.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")
    if adapt_df is not None and not adapt_df.empty:
        _write_df(ws_ad, adapt_df, start_row=4, start_col=1)
        for ri in range(5, 5 + len(adapt_df)):
            for ci in range(1, len(adapt_df.columns) + 1):
                ws_ad.cell(row=ri, column=ci).fill = BEST_FILL
                v = ws_ad.cell(row=ri, column=ci).value
                if isinstance(v, (float, int)) and not isinstance(v, bool):
                    ws_ad.cell(row=ri, column=ci).number_format = "0.0000"
        ws_ad.freeze_panes = "A5"
    else:
        ws_ad.cell(
            4,
            1,
            adapt_note or "Could not build adaptive selected table.",
        ).alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, 8):
        ws_ad.column_dimensions[get_column_letter(col)].width = 18

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
