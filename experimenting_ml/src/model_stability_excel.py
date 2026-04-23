"""
Per-target CV rank stability workbook: fold-level ranks and rank variance for a shortlist.

Uses aligned ``fold_rmses`` from ``cv_results.json`` (same as Friedman/Nemenyi).
Lower RMSE => rank 1 in each fold.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.stats import rankdata

from step2_ranking import (
    build_fold_rmse_matrix,
    load_cv_results,
    shortlist_by_median_mean_rank,
)


def shortlist_topk_mean_cv_rmse(
    cv: Dict[str, Any],
    target: str,
    model_names: List[str],
    k: int,
) -> List[str]:
    """Top ``k`` models by lowest mean CV RMSE for this target only."""
    k = max(1, int(k))
    scored = [(float(cv[target][m]["mean_rmse"]), m) for m in model_names]
    scored.sort(key=lambda x: (x[0], x[1]))
    return [m for _, m in scored[:k]]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=12, color="1F4E79")


def _safe_sheet_name(name: str) -> str:
    invalid = r"[]:*?/\\"
    s = "".join("_" if c in invalid else c for c in name)
    return s[:31] if len(s) > 31 else s


def _per_fold_rank_matrix(rmse_mat: pd.DataFrame) -> pd.DataFrame:
    vals = rmse_mat.values.astype(float)
    ranks = np.zeros_like(vals, dtype=float)
    for i in range(len(rmse_mat)):
        ranks[i, :] = rankdata(vals[i, :], method="average")
    return pd.DataFrame(ranks, index=rmse_mat.index, columns=rmse_mat.columns)


def _rank_stability_table(
    rank_mat: pd.DataFrame,
    cv_target: Dict[str, Any],
    models: List[str],
) -> pd.DataFrame:
    rows = []
    n_f = len(rank_mat)
    for m in models:
        r = cv_target[m]
        rr = rank_mat[m].values.astype(float)
        rows.append(
            {
                "model": m,
                "mean_cv_rmse": float(r["mean_rmse"]),
                "std_cv_rmse": float(r["std_rmse"]),
                "mean_fold_rank": float(np.mean(rr)),
                "std_fold_rank": float(np.std(rr, ddof=1)) if n_f > 1 else 0.0,
                "min_fold_rank": float(np.min(rr)),
                "max_fold_rank": float(np.max(rr)),
                "fold_rank_range": float(np.max(rr) - np.min(rr)),
            }
        )
    df = pd.DataFrame(rows)
    df = df.sort_values(
        ["std_fold_rank", "mean_fold_rank", "mean_cv_rmse"]
    ).reset_index(drop=True)
    df["stability_order"] = range(1, len(df) + 1)
    return df


def _write_df_block(
    ws,
    df: pd.DataFrame,
    start_row: int,
    *,
    header: bool = True,
) -> int:
    r = start_row
    cols = list(df.columns)
    if header:
        for ci, c in enumerate(cols, 1):
            cell = ws.cell(r, ci, c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        r += 1
    for _, row in df.iterrows():
        for ci, c in enumerate(cols, 1):
            v = row[c]
            ws.cell(r, ci, float(v) if isinstance(v, (float, np.floating)) else v)
        r += 1
    return r


def _write_rank_matrix(ws, rank_mat: pd.DataFrame, start_row: int) -> None:
    r = start_row
    ws.cell(r, 1, "fold").font = HEADER_FONT
    ws.cell(r, 1).fill = HEADER_FILL
    for j, m in enumerate(rank_mat.columns, 2):
        c = ws.cell(r, j, m)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r += 1
    for idx, row in rank_mat.iterrows():
        ws.cell(r, 1, int(idx) if not isinstance(idx, tuple) else idx)
        for j, m in enumerate(rank_mat.columns, 2):
            ws.cell(r, j, float(row[m])).number_format = "0.00"
        r += 1


def build_model_stability_workbook(
    out_path: Path,
    *,
    cv_path: Path,
    targets: List[str],
    model_names: List[str],
    shortlist: List[str],
    shortlist_mode: str = "topk_mean_cv_rmse",
    shortlist_k: int = 5,
) -> None:
    cv = load_cv_results(cv_path)
    global_sl = shortlist
    if shortlist_mode == "global_median_rank":
        pass  # use shortlist as global list
    elif shortlist_mode == "topk_mean_cv_rmse":
        global_sl = []  # unused; per-target list computed in loop
    elif shortlist_mode == "all_models":
        global_sl = list(model_names)
    wb = Workbook()
    wb.remove(wb.active)

    wm = wb.create_sheet("Methodology", 0)
    wm.cell(1, 1, "Model stability validation — how to read this workbook").font = TITLE_FONT
    lines = [
        "",
        "Purpose: show whether shortlisted models keep a stable *relative* ranking across CV folds.",
        "",
        "Shortlist modes (see run_model_stability_excel.py --shortlist-mode):",
        "  • topk_mean_cv_rmse (recommended): per target, the k models with lowest mean CV RMSE — stability among real contenders.",
        "  • global_median_rank: same k models for every target (median of mean fold ranks across targets; aligns with Step 2 shortlist).",
        "  • all_models: rank across the full zoo (wide tables).",
        "",
        "Per fold: models are ranked by validation RMSE (1 = best). Ranks are only comparable within a target sheet (shortlist).",
        "",
        "Key columns:",
        "  • std_fold_rank — standard deviation of fold ranks across CV scores; lower = more stable ranking.",
        "  • std_cv_rmse — standard deviation of fold RMSEs; lower = more stable absolute error.",
        "  • mean_fold_rank — average rank across folds (lower = better on average).",
        "",
        "Data source: cv_results.json fold_rmses (aligned fold order across models).",
    ]
    for i, line in enumerate(lines, start=2):
        wm.cell(i, 1, line)
        wm.cell(i, 1).alignment = Alignment(wrap_text=True, vertical="top")
    wm.column_dimensions["A"].width = 92

    summary_rows: List[Dict[str, Any]] = []

    for target in targets:
        rmse_mat = build_fold_rmse_matrix(cv, target, model_names)
        if shortlist_mode == "all_models":
            short = list(model_names)
        elif shortlist_mode == "topk_mean_cv_rmse":
            short = shortlist_topk_mean_cv_rmse(
                cv, target, model_names, shortlist_k
            )
        else:
            short = [m for m in global_sl if m in rmse_mat.columns]
            if not short:
                short = [model_names[0]]
        rmse_sub = rmse_mat[short]
        rank_mat = _per_fold_rank_matrix(rmse_sub)
        stab = _rank_stability_table(rank_mat, cv[target], short)

        for _, row in stab.iterrows():
            summary_rows.append(
                {
                    "target": target,
                    "model": row["model"],
                    "mean_cv_rmse": row["mean_cv_rmse"],
                    "std_cv_rmse": row["std_cv_rmse"],
                    "mean_fold_rank": row["mean_fold_rank"],
                    "std_fold_rank": row["std_fold_rank"],
                    "stability_order": int(row["stability_order"]),
                }
            )

        st = _safe_sheet_name(target)
        ws = wb.create_sheet(title=st)
        ws.cell(1, 1, f"CV rank stability — {target}").font = TITLE_FONT
        if shortlist_mode == "topk_mean_cv_rmse":
            sl_note = f"Top {shortlist_k} by mean CV RMSE (this target only):"
        elif shortlist_mode == "all_models":
            sl_note = "All models:"
        else:
            sl_note = "Shortlist (global median mean-rank across targets):"
        ws.cell(2, 1, sl_note)
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=6)
        ws.cell(2, 2, "; ".join(short))
        ws.cell(2, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(3, 1, "n_cv_scores (folds):")
        ws.cell(3, 2, len(rmse_sub))
        ws.cell(4, 1, "Per-fold ranks: 1 = best RMSE among shortlist in that fold.")

        r0 = 6
        ws.cell(r0, 1, "Rank stability (lower std_fold_rank = more stable ranking across folds)").font = Font(
            bold=True
        )
        _write_df_block(ws, stab, r0 + 1)

        r1 = r0 + 2 + len(stab) + 1
        ws.cell(r1, 1, "Fold × model rank matrix").font = Font(bold=True)
        _write_rank_matrix(ws, rank_mat, r1 + 1)

        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = min(18, 48)

    ws0 = wb.create_sheet("Summary", 1)
    ws0.cell(1, 1, "Shortlist rank stability across all targets").font = TITLE_FONT
    ws0.cell(
        2,
        1,
        "stability_order: 1 = most stable (lowest std_fold_rank) within that target shortlist.",
    ).font = Font(italic=True)
    if summary_rows:
        sdf = pd.DataFrame(summary_rows).sort_values(
            ["target", "std_fold_rank", "mean_fold_rank"]
        )
        _write_df_block(ws0, sdf, 4)
    for col in range(1, 10):
        ws0.column_dimensions[get_column_letter(col)].width = 16

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def compute_global_median_shortlist(
    cv_path: Path,
    model_names: List[str],
    targets: List[str],
    *,
    shortlist_k: int,
) -> List[str]:
    return shortlist_by_median_mean_rank(
        cv_path, model_names, targets, k=shortlist_k
    )
