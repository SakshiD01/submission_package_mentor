"""
Join SHAP *__importance.csv outputs with docs/nolhc_inputs_crosswalk.csv for XAI review.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import List, Optional, Tuple

import pandas as pd


def _safe_filename(s: str) -> str:
    return re.sub(r"[^a-zA-Z0-9_.-]+", "_", s)


def _parse_importance_path(path: Path) -> Optional[Tuple[str, str]]:
    """
    Filenames: ``<target_safe>__<model_safe>__importance.csv`` (from run_step4_shap).
    """
    name = path.name
    if not name.endswith("__importance.csv"):
        return None
    base = name[: -len("__importance.csv")]
    if "__" not in base:
        return None
    target_s, model_s = base.rsplit("__", 1)
    return target_s, model_s


def load_crosswalk(crosswalk_csv: Path) -> pd.DataFrame:
    df = pd.read_csv(crosswalk_csv)
    if "feature_name" not in df.columns:
        raise ValueError(f"{crosswalk_csv} must have column feature_name")
    return df


def build_shap_review_long_table(
    shap_dir: Path,
    crosswalk_csv: Path,
    *,
    top_k: int = 5,
    selected_models_csv: Optional[Path] = None,
) -> pd.DataFrame:
    """
    For each (target, model) importance file, take top ``top_k`` features by mean_abs_shap
    and left-merge crosswalk descriptions.
    """
    cw = load_crosswalk(crosswalk_csv)
    cw = cw.rename(columns=lambda c: c.strip())

    paths: List[Path] = []
    if selected_models_csv is not None and selected_models_csv.is_file():
        sel = pd.read_csv(selected_models_csv)
        for _, row in sel.iterrows():
            t = str(row["target"])
            m = str(row["model"])
            p = shap_dir / f"{_safe_filename(t)}__{_safe_filename(m)}__importance.csv"
            if p.is_file():
                paths.append(p)
    else:
        paths = sorted(shap_dir.glob("*__importance.csv"))

    out_rows = []
    for p in paths:
        parsed = _parse_importance_path(p)
        if parsed is None:
            continue
        target_s, model_s = parsed
        imp = pd.read_csv(p)
        if "feature" not in imp.columns or "mean_abs_shap" not in imp.columns:
            continue
        sub = imp.sort_values("mean_abs_shap", ascending=False).head(max(1, top_k)).copy()
        sub["shap_rank"] = range(1, len(sub) + 1)
        sub["target_key"] = target_s
        sub["model_key"] = model_s
        sub["importance_file"] = p.name
        merged = sub.merge(
            cw,
            left_on="feature",
            right_on="feature_name",
            how="left",
        )
        merged = merged.drop(columns=["feature_name"], errors="ignore")
        if "short_description" in merged.columns:
            missing = merged["short_description"].isna() | (
                merged["short_description"].astype(str).str.strip() == ""
            )
            merged.loc[missing, "short_description"] = merged.loc[missing, "feature"].apply(
                lambda x: f"(add to crosswalk) {x}"
            )
        else:
            merged["short_description"] = merged["feature"].apply(
                lambda x: f"(add to crosswalk) {x}"
            )
        out_rows.append(merged)

    if not out_rows:
        return pd.DataFrame()
    df = pd.concat(out_rows, ignore_index=True)
    preferred = [
        "target_key",
        "model_key",
        "shap_rank",
        "feature",
        "mean_abs_shap",
        "short_description",
        "unit",
        "sig_category",
        "expected_direction_notes",
        "xai_review_expected_surprising_flag",
        "importance_file",
    ]
    cols = [c for c in preferred if c in df.columns] + [
        c for c in df.columns if c not in preferred
    ]
    return df[cols]


def write_shap_xai_review_artifacts(
    out_dir: Path,
    long_df: pd.DataFrame,
) -> Optional[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    long_path = out_dir / "xai_shap_top_features_long.csv"
    long_df.to_csv(long_path, index=False)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Top_features_x_domain"
    if long_df.empty:
        ws.cell(1, 1, "(no SHAP importance files found)")
    else:
        for c, col in enumerate(long_df.columns, 1):
            ws.cell(1, c, col).font = Font(bold=True)
        for ri, row in enumerate(long_df.itertuples(index=False), 2):
            for ci, val in enumerate(row, 1):
                ws.cell(ri, ci, val if val is not None and not (isinstance(val, float) and pd.isna(val)) else "")
        ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Review_checklist")
    lines = [
        "XAI attribution review (before stakeholder-facing text)",
        "",
        "1. For each target row block: confirm top features align with SIG/technical-report semantics.",
        "2. Mark surprising drivers: data artefact vs plausible mechanism.",
        "3. If train vs test SHAP was run: note any rank flips in a sentence.",
        "4. Fill xai_review_expected_surprising_flag in docs/nolhc_inputs_crosswalk.csv if needed.",
        "",
        "This workbook is generated; edit the CSV crosswalk for persistent notes.",
    ]
    for i, line in enumerate(lines, 1):
        ws2.cell(i, 1, line)
        ws2.cell(i, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.column_dimensions["A"].width = 92

    xlsx_path = out_dir / "xai_shap_domain_review.xlsx"
    wb.save(xlsx_path)
    return xlsx_path
