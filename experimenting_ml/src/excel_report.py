"""
Step 10: formatted Excel workbook (ML_Pipeline_Specification.md §10, §3).
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
SIG_FILL = PatternFill("solid", fgColor="FFF2CC")
BEST_FILL = PatternFill("solid", fgColor="E2EFDA")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
NORMAL_FONT = Font(name="Arial", size=10)
RED_BOLD = Font(name="Arial", size=9, bold=True, color="FF0000")
SMALL_FONT = Font(name="Arial", size=9)
TITLE_FONT = Font(bold=True, name="Arial", size=12, color="1F4E79")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Lower = preferred on tie-break (spec §3.3)
SIMPLICITY_RANK: Dict[str, int] = {
    "Ridge": 0,
    "Lasso": 1,
    "ElasticNet": 2,
    "BayesianRidge": 3,
    "PolynomialReg_deg2": 4,
    "PolynomialReg_deg3": 5,
    "KNN": 6,
    "MLP": 7,
    "SVR_Poly": 8,
    "SVR_RBF": 9,
    "AdaBoost": 10,
    "GradientBoosting": 11,
    "RandomForest": 12,
    "ExtraTrees": 13,
    "XGBoost": 14,
    "LightGBM": 15,
    "CatBoost": 16,
    "GPR_RBF": 17,
    "GPR_Matern": 18,
    "Baseline_Mean": -2,
    "Baseline_OLS": -1,
}


def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    s = re.sub(r"[\[\]\*\/\\?:]", "_", name)
    return s[:max_len]


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def build_ttest_structures(
    pairs_csv: Path, model_names: List[str], targets: List[str]
) -> Dict[str, Dict[str, Any]]:
    """pairs_df + symmetric p-value matrix per target."""
    df = pd.read_csv(pairs_csv)
    out: Dict[str, Dict[str, Any]] = {}
    for target in targets:
        sub = df[df["target"] == target].copy()
        cols = [
            "Model_A",
            "Model_B",
            "Mean_RMSE_A",
            "Mean_RMSE_B",
            "t_stat",
            "p_value",
            "Significant",
            "Better_Model",
        ]
        pairs_df = sub[cols].copy()
        mat = pd.DataFrame(np.nan, index=model_names, columns=model_names, dtype=float)
        for _, row in sub.iterrows():
            a, b = row["Model_A"], row["Model_B"]
            p = float(row["p_value"])
            mat.loc[a, b] = p
            mat.loc[b, a] = p
        out[target] = {"pairs_df": pairs_df, "matrix_df": mat}
    return out


def count_ttest_wins(pairs_df: pd.DataFrame, model: str) -> int:
    w = 0
    for _, row in pairs_df.iterrows():
        if not bool(row["Significant"]):
            continue
        bm = row["Better_Model"]
        if pd.isna(bm) or bm == "":
            continue
        if str(bm) == model:
            w += 1
    return w


def compute_model_selection(
    cv_results: Dict[str, Any],
    test_results: Dict[str, Any],
    conformal_results: Dict[str, Any],
    ttest_by_target: Dict[str, Dict[str, Any]],
    targets: List[str],
    model_names: List[str],
) -> pd.DataFrame:
    rows = []
    eps = 1e-10

    for target in targets:
        cv_means = {m: cv_results[target][m]["mean_rmse"] for m in model_names}
        test_rmses = {m: test_results[target][m]["rmse"] for m in model_names}
        min_cv, max_cv = min(cv_means.values()), max(cv_means.values())
        min_te, max_te = min(test_rmses.values()), max(test_rmses.values())

        pairs_df = ttest_by_target[target]["pairs_df"]
        wins = {m: count_ttest_wins(pairs_df, m) for m in model_names}
        n_peer = max(1, len(model_names) - 1)

        composite = {}
        for m in model_names:
            ncv = (cv_means[m] - min_cv) / (max_cv - min_cv + eps)
            nte = (test_rmses[m] - min_te) / (max_te - min_te + eps)
            composite[m] = (
                0.4 * (1.0 - ncv)
                + 0.4 * (1.0 - nte)
                + 0.2 * (wins[m] / float(n_peer))
            )

        best_score = max(composite.values())
        cand = [m for m in model_names if best_score - composite[m] <= 0.01 + 1e-12]

        def tie_key(m: str) -> Tuple[float, int]:
            std_cv = cv_results[target][m]["std_rmse"]
            return (std_cv, SIMPLICITY_RANK.get(m, 99))

        best_m = min(cand, key=tie_key)

        cv_r = cv_results[target][best_m]
        te_r = test_results[target][best_m]
        cf_r = conformal_results[target][best_m]
        just = (
            f"Highest composite score ({composite[best_m]:.4f}); "
            f"CV RMSE={cv_r['mean_rmse']:.4f} (std {cv_r['std_rmse']:.4f}); "
            f"test RMSE={te_r['rmse']:.4f}, R²={te_r['r2']:.4f}; "
            f"{wins[best_m]}/{n_peer} possible significant CV-fold wins vs other models."
        )

        rows.append(
            {
                "Target": target,
                "Best_Model": best_m,
                "CV_Mean_RMSE": cv_r["mean_rmse"],
                "CV_Std_RMSE": cv_r["std_rmse"],
                "Test_RMSE": te_r["rmse"],
                "Test_MAE": te_r["mae"],
                "Test_R2": te_r["r2"],
                "TTest_Wins": wins[best_m],
                "Composite_Score": composite[best_m],
                "Conformal_Coverage": cf_r["coverage_level"],
                "Empirical_Coverage": cf_r["empirical_coverage"],
                "Interval_Width": cf_r["interval_width"],
                "Justification": just,
            }
        )

    return pd.DataFrame(rows)


def compute_model_selection_pre_test(
    cv_results: Dict[str, Any],
    ttest_by_target: Dict[str, Dict[str, Any]],
    targets: List[str],
    model_names: List[str],
) -> pd.DataFrame:
    """
    Same tie-breaking as ``compute_model_selection``, but composite uses only
    **CV** and **paired t-test wins** (no hold-out test metrics).

    Weights mirror the full composite's CV:wins ratio (0.4 : 0.2 → 2/3 vs 1/3)
    so this is the natural pre–test-evaluation analogue. Use before conformal / SHAP
    when ``test_results.json`` does not exist yet.
    """
    rows = []
    eps = 1e-10

    for target in targets:
        cv_means = {m: cv_results[target][m]["mean_rmse"] for m in model_names}
        min_cv, max_cv = min(cv_means.values()), max(cv_means.values())

        pairs_df = ttest_by_target[target]["pairs_df"]
        wins = {m: count_ttest_wins(pairs_df, m) for m in model_names}
        n_peer = max(1, len(model_names) - 1)

        composite = {}
        for m in model_names:
            ncv = (cv_means[m] - min_cv) / (max_cv - min_cv + eps)
            composite[m] = (2.0 / 3.0) * (1.0 - ncv) + (1.0 / 3.0) * (
                wins[m] / float(n_peer)
            )

        best_score = max(composite.values())
        cand = [m for m in model_names if best_score - composite[m] <= 0.01 + 1e-12]

        def tie_key(m: str) -> Tuple[float, int]:
            std_cv = cv_results[target][m]["std_rmse"]
            return (std_cv, SIMPLICITY_RANK.get(m, 99))

        best_m = min(cand, key=tie_key)
        cv_r = cv_results[target][best_m]
        just = (
            f"Pre-test composite ({composite[best_m]:.4f}): 2/3·(1-norm CV RMSE) + 1/3·(t-test wins/{n_peer}); "
            f"CV RMSE={cv_r['mean_rmse']:.4f} (std {cv_r['std_rmse']:.4f}); "
            f"{wins[best_m]} significant CV-fold wins. "
            f"(Full Step 10 composite adds test RMSE when available.)"
        )

        rows.append(
            {
                "Target": target,
                "Best_Model": best_m,
                "CV_Mean_RMSE": cv_r["mean_rmse"],
                "CV_Std_RMSE": cv_r["std_rmse"],
                "TTest_Wins": wins[best_m],
                "Composite_PreTest_Score": composite[best_m],
                "Justification": just,
            }
        )

    return pd.DataFrame(rows)


def style_header(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def write_df(
    ws,
    df: pd.DataFrame,
    start_row: int = 1,
    *,
    highlight_sig: bool = False,
    sig_col: str = "Significant",
) -> int:
    """Write dataframe; return last row index used."""
    ncols = len(df.columns)
    for ci, col in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=ci, value=col)
    style_header(ws, start_row, ncols)

    last = start_row
    for ri, (_, row_data) in enumerate(df.iterrows(), start_row + 1):
        last = ri
        sig = False
        if highlight_sig and sig_col in df.columns:
            v = row_data[sig_col]
            sig = bool(v) if not isinstance(v, str) else str(v).lower() == "true"
        if highlight_sig and sig:
            row_fill = SIG_FILL
        else:
            row_fill = ALT_FILL if ri % 2 == 0 else PatternFill()

        for ci, col in enumerate(df.columns, 1):
            val = row_data[col]
            if isinstance(val, (float, np.floating)) and (np.isnan(val) or np.isinf(val)):
                val = None
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = NORMAL_FONT
            cell.alignment = CENTER
            cell.fill = row_fill
            if isinstance(val, bool):
                pass
            elif isinstance(val, (float, np.floating, int)):
                cell.number_format = "0.0000"

    return last


def write_merged_title(ws, row: int, text: str, last_col: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, name="Arial", size=11, color="1F4E79")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_mini_best_model_block(
    ws,
    start_row: int,
    selection_df: pd.DataFrame,
    targets: List[str],
    *,
    last_col_merge: int = 12,
) -> int:
    """
    Write title + compact table of selected best model metrics for all targets.
    Returns the row index for the main sheet's column header row.
    """
    write_merged_title(
        ws,
        start_row,
        "Selected best model per target (composite score: 40% CV RMSE + 40% test RMSE + 20% t-test wins / 18)",
        last_col_merge,
    )
    mini_cols = [
        "Target",
        "Best_Model",
        "CV_Mean_RMSE",
        "CV_Std_RMSE",
        "Test_RMSE",
        "Test_MAE",
        "Test_R2",
        "TTest_Wins",
        "Composite_Score",
        "Conformal_Coverage",
        "Empirical_Coverage",
        "Interval_Width",
    ]
    sub = selection_df.set_index("Target").reindex(targets).reset_index()
    sub = sub[mini_cols]
    hdr_row = start_row + 1
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for ci, name in enumerate(mini_cols, 1):
        cell = ws.cell(row=hdr_row, column=ci, value=name)
        cell.fill = HEADER_FILL
        cell.font = hdr_font
        cell.alignment = CENTER
    data0 = hdr_row + 1
    for ri, (_, r) in enumerate(sub.iterrows(), data0):
        fill = BEST_FILL if ri % 2 == 1 else PatternFill()
        for ci, col in enumerate(mini_cols, 1):
            val = r[col]
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = NORMAL_FONT
            cell.alignment = CENTER
            cell.fill = fill
            if isinstance(val, bool):
                pass
            elif isinstance(val, (float, np.floating, int)):
                cell.number_format = "0.0000"
    return data0 + len(sub) + 1  # main table header row (after one blank row)


def highlight_rows_where_yes(
    ws,
    start_data_row: int,
    n_rows: int,
    yes_col_index: int,
) -> None:
    """yes_col_index is 1-based Excel column index for 'Is_Selected_Best_Model'."""
    for ri in range(start_data_row, start_data_row + n_rows):
        v = ws.cell(row=ri, column=yes_col_index).value
        if v == "Yes":
            ncols = ws.max_column
            for ci in range(1, ncols + 1):
                ws.cell(row=ri, column=ci).fill = BEST_FILL


def _target_selection_banner(target: str, selection_df: pd.DataFrame) -> str:
    r = selection_df.loc[selection_df["Target"] == target].iloc[0]
    return (
        f"Target: {target}  |  SELECTED BEST MODEL: {r['Best_Model']}  |  "
        f"CV RMSE: {r['CV_Mean_RMSE']:.4f} (std {r['CV_Std_RMSE']:.4f})  |  "
        f"Test RMSE: {r['Test_RMSE']:.4f}  |  Test MAE: {r['Test_MAE']:.4f}  |  "
        f"R²: {r['Test_R2']:.4f}  |  Composite: {r['Composite_Score']:.4f}  |  "
        f"T-test wins: {int(r['TTest_Wins'])}/18  |  "
        f"Conformal coverage: {r['Conformal_Coverage']}  |  "
        f"Empirical coverage: {r['Empirical_Coverage']:.4f}  |  "
        f"Interval width: {r['Interval_Width']:.4f}"
    )


def autofit_columns(ws, max_width: float = 40.0) -> None:
    for col in ws.columns:
        col0 = col[0].column
        letter = get_column_letter(col0)
        max_len = max((len(str(c.value)) if c.value is not None else 0 for c in col), default=0)
        ws.column_dimensions[letter].width = min(max_len + 4, max_width)


def _mentor_steps_rows(experiment_status: Dict[str, bool]) -> List[Dict[str, str]]:
    def present(rel_path: str) -> str:
        return "Present" if experiment_status.get(rel_path, False) else "Missing"

    rows = [
        {
            "Step": "Step 1 - CV Hyperparameter Search",
            "Purpose": "Train-split model tuning and fold-level diagnostics.",
            "Expected_Artifact": "outputs/cv_results.json",
            "Status": present("outputs/cv_results.json"),
        },
        {
            "Step": "Step 1 - Fold Details (extra diagnostics)",
            "Purpose": "Per-fold/per-parameter audit trail for stability and HP sensitivity.",
            "Expected_Artifact": "outputs/cv_fold_details.csv",
            "Status": present("outputs/cv_fold_details.csv"),
        },
        {
            "Step": "Paired T-tests",
            "Purpose": "Significance checks for pairwise CV performance differences.",
            "Expected_Artifact": "outputs/paired_ttests_all_targets.csv",
            "Status": present("outputs/paired_ttests_all_targets.csv"),
        },
        {
            "Step": "Mentor Step 2 (Friedman/Nemenyi + figures)",
            "Purpose": "Global multi-model comparison plus CD/learning-curve/residual visuals.",
            "Expected_Artifact": "outputs/step2/friedman_nemenyi_summary.csv",
            "Status": present("outputs/step2/friedman_nemenyi_summary.csv"),
        },
        {
            "Step": "Mentor Step 2 Master Workbook",
            "Purpose": "One-sheet-per-target Step 2 evidence pack.",
            "Expected_Artifact": "outputs/experiment_master.xlsx",
            "Status": present("outputs/experiment_master.xlsx"),
        },
        {
            "Step": "Mentor Step 3 (pre-conformal)",
            "Purpose": "Selection, baselines, calibration, and HP sensitivity before test/conformal.",
            "Expected_Artifact": "outputs/step3_pre_conformal",
            "Status": present("outputs/step3_pre_conformal"),
        },
        {
            "Step": "Step 6 - Retrain Selected Models",
            "Purpose": "Refit chosen models on training split for downstream evaluation/XAI.",
            "Expected_Artifact": "outputs/trained_models/split_meta.json",
            "Status": present("outputs/trained_models/split_meta.json"),
        },
        {
            "Step": "Step 4 - SHAP XAI",
            "Purpose": "Feature attribution for selected per-target models.",
            "Expected_Artifact": "outputs/step4_shap/shap_selected_models.csv",
            "Status": present("outputs/step4_shap/shap_selected_models.csv"),
        },
        {
            "Step": "Step 4 - SHAP Master Workbook",
            "Purpose": "One-tab-per-target SHAP summary workbook.",
            "Expected_Artifact": "outputs/shap_master.xlsx",
            "Status": present("outputs/shap_master.xlsx"),
        },
        {
            "Step": "Steps 7 and 9 - Test + Conformal",
            "Purpose": "Holdout performance and uncertainty interval metrics.",
            "Expected_Artifact": "outputs/test_results.json",
            "Status": present("outputs/test_results.json"),
        },
        {
            "Step": "Step 10 - Final Integrated Workbook",
            "Purpose": "Consolidated report with selection, tests, conformal, and mentor log.",
            "Expected_Artifact": "outputs/pipeline_results.xlsx",
            "Status": present("outputs/pipeline_results.xlsx"),
        },
    ]
    return rows


def generate_excel_report(
    *,
    cv_results: Dict[str, Any],
    test_results: Dict[str, Any],
    conformal_results: Dict[str, Any],
    ttest_by_target: Dict[str, Dict[str, Any]],
    targets: List[str],
    model_names: List[str],
    output_path: Path,
    experiment_status: Optional[Dict[str, bool]] = None,
) -> None:
    selection_df = compute_model_selection(
        cv_results,
        test_results,
        conformal_results,
        ttest_by_target,
        targets,
        model_names,
    )

    wb = Workbook()
    wb.remove(wb.active)

    sel_map = selection_df.set_index("Target")["Best_Model"].to_dict()

    # --- Summary ---
    ws = wb.create_sheet("Summary", 0)
    write_merged_title(
        ws,
        1,
        "Executive summary — selected surrogate model per target (ranked by test RMSE on 26 holdout runs)",
        12,
    )
    summary_cols = [
        "Target",
        "Best_Model",
        "CV_Mean_RMSE",
        "CV_Std_RMSE",
        "Test_RMSE",
        "Test_MAE",
        "Test_R2",
        "TTest_Wins",
        "Composite_Score",
        "Conformal_Coverage",
        "Empirical_Coverage",
        "Interval_Width",
    ]
    summary_df = selection_df[summary_cols].sort_values("Test_RMSE").reset_index(drop=True)
    write_df(ws, summary_df, start_row=2)
    for ri in range(3, 3 + len(summary_df)):
        for ci in range(1, len(summary_cols) + 1):
            ws.cell(row=ri, column=ci).fill = BEST_FILL
            v = ws.cell(row=ri, column=ci).value
            if isinstance(v, bool):
                pass
            elif isinstance(v, (float, np.floating, int)):
                ws.cell(row=ri, column=ci).number_format = "0.0000"
    ws.freeze_panes = "A3"
    autofit_columns(ws)

    # --- CV_Results ---
    ws = wb.create_sheet("CV_Results")
    main_hdr = write_mini_best_model_block(ws, 1, selection_df, targets)
    cv_rows = []
    for target in targets:
        best = sel_map[target]
        for model in model_names:
            r = cv_results[target][model]
            cv_rows.append(
                {
                    "Target": target,
                    "Model": model,
                    "Selected_Best_Model": best,
                    "Is_Selected_Best_Model": "Yes" if model == best else "No",
                    "Best_Params": str(r["best_params"]),
                    "Mean_RMSE": r["mean_rmse"],
                    "Std_RMSE": r["std_rmse"],
                }
            )
    cv_df = pd.DataFrame(cv_rows)
    write_df(ws, cv_df, start_row=main_hdr)
    yes_col = list(cv_df.columns).index("Is_Selected_Best_Model") + 1
    highlight_rows_where_yes(ws, main_hdr + 1, len(cv_rows), yes_col)
    ws.freeze_panes = f"A{main_hdr + 1}"
    autofit_columns(ws)

    # --- Test_Results ---
    ws = wb.create_sheet("Test_Results")
    main_hdr = write_mini_best_model_block(ws, 1, selection_df, targets)
    test_rows = []
    for target in targets:
        best = sel_map[target]
        for model in model_names:
            r = test_results[target][model]
            test_rows.append(
                {
                    "Target": target,
                    "Model": model,
                    "Selected_Best_Model": best,
                    "Is_Selected_Best_Model": "Yes" if model == best else "No",
                    "RMSE": r["rmse"],
                    "MAE": r["mae"],
                    "R2": r["r2"],
                }
            )
    test_df = pd.DataFrame(test_rows)
    write_df(ws, test_df, start_row=main_hdr)
    yes_col = list(test_df.columns).index("Is_Selected_Best_Model") + 1
    highlight_rows_where_yes(ws, main_hdr + 1, len(test_rows), yes_col)
    n_models = len(model_names)
    rmse_col_letter = get_column_letter(list(test_df.columns).index("RMSE") + 1)
    for ti, _target in enumerate(targets):
        r0 = main_hdr + 1 + ti * n_models
        r1 = r0 + n_models - 1
        ws.conditional_formatting.add(
            f"{rmse_col_letter}{r0}:{rmse_col_letter}{r1}",
            ColorScaleRule(
                start_type="min",
                start_color="63BE7B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="F8696B",
            ),
        )
    ws.freeze_panes = f"A{main_hdr + 1}"
    autofit_columns(ws)

    # --- Per-target t-test sheets ---
    n_pair = len(ttest_by_target[targets[0]]["pairs_df"])
    for ti, target in enumerate(targets):
        sheet_name = _safe_sheet_name(f"Target_{ti + 1:02d}_{target}")
        ws = wb.create_sheet(sheet_name)
        write_merged_title(ws, 1, _target_selection_banner(target, selection_df), 12)
        ws.cell(row=2, column=1, value="PAIRWISE PAIRED T-TEST RESULTS")
        ws.cell(row=2, column=1).font = TITLE_FONT

        pairs_df = ttest_by_target[target]["pairs_df"]
        write_df(ws, pairs_df, start_row=3, highlight_sig=True)
        offset = 3 + n_pair + 3
        ws.cell(row=offset, column=1, value="P-VALUE MATRIX (19x19)")
        ws.cell(row=offset, column=1).font = TITLE_FONT

        matrix_df = ttest_by_target[target]["matrix_df"]
        hdr = offset + 1
        hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        for ci, colm in enumerate(model_names, 2):
            c = ws.cell(row=hdr, column=ci, value=colm)
            c.fill = HEADER_FILL
            c.font = hdr_font
            c.alignment = CENTER
        for ri, row_m in enumerate(model_names, hdr + 1):
            c = ws.cell(row=ri, column=1, value=row_m)
            c.fill = HEADER_FILL
            c.font = hdr_font
            c.alignment = CENTER
            for ci, col_m in enumerate(model_names, 2):
                val = matrix_df.loc[row_m, col_m]
                cell = ws.cell(row=ri, column=ci)
                if pd.isna(val):
                    cell.value = None
                else:
                    cell.value = float(val)
                    cell.number_format = "0.0000"
                    if val < 0.05:
                        cell.font = RED_BOLD
                    else:
                        cell.font = SMALL_FONT

        mat_start = hdr + 1
        mat_end = hdr + len(model_names)
        ws.conditional_formatting.add(
            f"B{mat_start}:T{mat_end}",
            ColorScaleRule(
                start_type="min",
                start_color="1F4E79",
                mid_type="num",
                mid_value=0.05,
                mid_color="FFFFFF",
                end_type="max",
                end_color="FFFFFF",
            ),
        )
        ws.freeze_panes = "A4"
        autofit_columns(ws)

    # --- Conformal ---
    ws = wb.create_sheet("Conformal_Prediction")
    main_hdr = write_mini_best_model_block(ws, 1, selection_df, targets)
    conf_rows = []
    for target in targets:
        best = sel_map[target]
        for model in model_names:
            c = conformal_results[target][model]
            conf_rows.append(
                {
                    "Target": target,
                    "Model": model,
                    "Selected_Best_Model": best,
                    "Is_Selected_Best_Model": "Yes" if model == best else "No",
                    "Coverage_Level": c["coverage_level"],
                    "Relative_RMSE_to_Best": c.get("relative_rmse_to_best", np.nan),
                    "Conformal_Quantile": c["quantile"],
                    "Empirical_Coverage": c["empirical_coverage"],
                    "Interval_Width": c["interval_width"],
                }
            )
    conf_df = pd.DataFrame(conf_rows)
    write_df(ws, conf_df, start_row=main_hdr)
    yes_col = list(conf_df.columns).index("Is_Selected_Best_Model") + 1
    highlight_rows_where_yes(ws, main_hdr + 1, len(conf_rows), yes_col)
    ws.freeze_panes = f"A{main_hdr + 1}"
    autofit_columns(ws)

    # --- Model_Selection ---
    ws = wb.create_sheet("Model_Selection")
    write_merged_title(
        ws,
        1,
        "Full model selection record (composite score + tie-breakers) — same best models as summary blocks on other sheets",
        12,
    )
    write_df(ws, selection_df, start_row=2)
    for ri in range(3, len(selection_df) + 3):
        for ci in range(1, len(selection_df.columns) + 1):
            ws.cell(row=ri, column=ci).fill = BEST_FILL
            ws.cell(row=ri, column=ci).font = NORMAL_FONT
            ws.cell(row=ri, column=ci).alignment = CENTER
            v = ws.cell(row=ri, column=ci).value
            if isinstance(v, bool):
                pass
            elif isinstance(v, (float, np.floating, int)):
                ws.cell(row=ri, column=ci).number_format = "0.0000"
    # Justification column: wider wrap
    just_col = list(selection_df.columns).index("Justification") + 1
    for ri in range(3, len(selection_df) + 3):
        ws.cell(row=ri, column=just_col).alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
    ws.column_dimensions[get_column_letter(just_col)].width = 60
    ws.freeze_panes = "A3"
    autofit_columns(ws)

    # --- Mentor / Extra experiments log ---
    ws = wb.create_sheet("Mentor_Steps_Log")
    write_merged_title(
        ws,
        1,
        "Mentor steps and extra experimentation log (auto-detected from outputs/)",
        4,
    )
    rows = _mentor_steps_rows(experiment_status or {})
    log_df = pd.DataFrame(rows)
    write_df(ws, log_df, start_row=2)
    for ri in range(3, len(log_df) + 3):
        status = ws.cell(row=ri, column=4).value
        fill = BEST_FILL if status == "Present" else ALT_FILL
        for ci in range(1, 5):
            ws.cell(row=ri, column=ci).fill = fill
            ws.cell(row=ri, column=ci).alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 14
    ws.freeze_panes = "A3"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
