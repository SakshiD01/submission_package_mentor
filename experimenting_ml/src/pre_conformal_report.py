"""
Build a single pre-conformal checkpoint workbook for mentor review.

This report intentionally excludes conformal/test-final sheets and focuses on:
Step 1 CV, paired t-tests, Step 2/Step 3 artifacts, and Step 4 SHAP selection.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(fill_type="solid", start_color="1F4E79", end_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
OK_FILL = PatternFill(fill_type="solid", start_color="E2F0D9", end_color="E2F0D9")
MISS_FILL = PatternFill(fill_type="solid", start_color="FCE4D6", end_color="FCE4D6")


def _write_df(ws, df: pd.DataFrame, start_row: int = 1) -> int:
    if df is None or df.empty:
        ws.cell(start_row, 1, "(no rows)")
        return start_row

    cols = list(df.columns)
    for ci, c in enumerate(cols, 1):
        cell = ws.cell(start_row, ci, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    r = start_row + 1
    for _, row in df.iterrows():
        for ci, c in enumerate(cols, 1):
            v = row[c]
            ws.cell(r, ci, "" if pd.isna(v) else v)
        r += 1
    return r - 1


def _autofit(ws, max_width: int = 56) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            s = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(s))
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def _read_csv_if_exists(path: Path) -> pd.DataFrame:
    if path.is_file():
        return pd.read_csv(path)
    return pd.DataFrame()


def _build_status_rows(out_dir: Path) -> List[Dict[str, str]]:
    checks = [
        ("Step 1 CV summary", out_dir / "cv_results.summary.csv"),
        ("Step 1 fold details", out_dir / "cv_fold_details.csv"),
        ("Paired t-tests", out_dir / "paired_ttests_all_targets.csv"),
        ("Step 2 summary", out_dir / "step2" / "friedman_nemenyi_summary.csv"),
        ("Step 2 master workbook", out_dir / "experiment_master.xlsx"),
        ("Step 3 selection", out_dir / "step3" / "per_target_cv_selection.csv"),
        ("Step 3 calibration", out_dir / "step3" / "calibration_cv_metrics.csv"),
        ("Step 3 HP sensitivity log", out_dir / "step3" / "hp_sensitivity_run_log.csv"),
        ("Step 3 workbook", out_dir / "step3_report.xlsx"),
        ("Step 4 SHAP selected models", out_dir / "step4_shap" / "shap_selected_models.csv"),
        ("Step 4 SHAP workbook", out_dir / "shap_master.xlsx"),
    ]
    rows: List[Dict[str, str]] = []
    for name, p in checks:
        rows.append(
            {
                "Artifact": name,
                "Path": str(p.relative_to(out_dir.parent)),
                "Status": "Present" if p.exists() else "Missing",
            }
        )
    return rows


def _build_shap_top_features(shap_dir: Path, selected_df: pd.DataFrame, top_n: int) -> pd.DataFrame:
    rows = []
    if selected_df.empty:
        return pd.DataFrame()
    for _, r in selected_df.iterrows():
        target = str(r["target"])
        model = str(r["model"])
        imp_path = shap_dir / f"{target}__{model}__importance.csv"
        if not imp_path.is_file():
            # fallback with sanitized names already produced by run_step4_shap
            from re import sub

            safe = lambda s: sub(r"[^a-zA-Z0-9_.-]+", "_", s)
            imp_path = shap_dir / f"{safe(target)}__{safe(model)}__importance.csv"
        if not imp_path.is_file():
            rows.append(
                {
                    "target": target,
                    "model": model,
                    "rank": "",
                    "feature": "(importance csv missing)",
                    "mean_abs_shap": "",
                }
            )
            continue
        imp = pd.read_csv(imp_path).head(max(1, top_n)).copy()
        imp["rank"] = range(1, len(imp) + 1)
        imp["target"] = target
        imp["model"] = model
        rows.extend(
            imp[["target", "model", "rank", "feature", "mean_abs_shap"]].to_dict(orient="records")
        )
    return pd.DataFrame(rows)


def build_pre_conformal_workbook(
    *,
    out_dir: Path,
    output_path: Path,
    shap_top_n: int = 10,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # 1) Overview
    ws = wb.create_sheet("Overview", 0)
    ws.cell(1, 1, "Pre-Conformal Checkpoint Report").font = Font(bold=True, size=14)
    ws.cell(2, 1, "Scope: mentor review before official holdout test + conformal.").font = Font(
        italic=True
    )
    status_df = pd.DataFrame(_build_status_rows(out_dir))
    _write_df(ws, status_df, start_row=4)
    for ri in range(5, 5 + len(status_df)):
        fill = OK_FILL if ws.cell(ri, 3).value == "Present" else MISS_FILL
        for ci in range(1, 4):
            ws.cell(ri, ci).fill = fill
    ws.freeze_panes = "A5"
    _autofit(ws)

    # 2) CV summary
    ws = wb.create_sheet("CV_Summary")
    cv_summary = _read_csv_if_exists(out_dir / "cv_results.summary.csv")
    if not cv_summary.empty:
        cv_summary = cv_summary.sort_values(["target", "mean_rmse"]).reset_index(drop=True)
    _write_df(ws, cv_summary, start_row=1)
    ws.freeze_panes = "A2"
    _autofit(ws)

    # 3) Paired t-tests
    ws = wb.create_sheet("Paired_TTests")
    ttests = _read_csv_if_exists(out_dir / "paired_ttests_all_targets.csv")
    _write_df(ws, ttests, start_row=1)
    ws.freeze_panes = "A2"
    _autofit(ws)

    # 4) Step 3 selection
    ws = wb.create_sheet("Step3_Selection")
    step3_sel = _read_csv_if_exists(out_dir / "step3" / "per_target_cv_selection.csv")
    _write_df(ws, step3_sel, start_row=1)
    ws.freeze_panes = "A2"
    _autofit(ws)

    # 5) Step 3 calibration
    ws = wb.create_sheet("Step3_Calibration")
    cal = _read_csv_if_exists(out_dir / "step3" / "calibration_cv_metrics.csv")
    _write_df(ws, cal, start_row=1)
    ws.freeze_panes = "A2"
    _autofit(ws)

    # 6) Step 3 HP sensitivity
    ws = wb.create_sheet("Step3_HP_Sensitivity")
    hp = _read_csv_if_exists(out_dir / "step3" / "hp_sensitivity_run_log.csv")
    _write_df(ws, hp, start_row=1)
    ws.freeze_panes = "A2"
    _autofit(ws)

    # 7) SHAP selected models
    ws = wb.create_sheet("SHAP_Selected_Models")
    shap_sel = _read_csv_if_exists(out_dir / "step4_shap" / "shap_selected_models.csv")
    _write_df(ws, shap_sel, start_row=1)
    ws.freeze_panes = "A2"
    _autofit(ws)

    # 8) SHAP top features per target
    ws = wb.create_sheet("SHAP_Top_Features")
    shap_top = _build_shap_top_features(out_dir / "step4_shap", shap_sel, top_n=shap_top_n)
    _write_df(ws, shap_top, start_row=1)
    ws.freeze_panes = "A2"
    _autofit(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

