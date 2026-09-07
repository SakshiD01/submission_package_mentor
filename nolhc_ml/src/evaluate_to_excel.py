"""
evaluate_to_excel.py — bundle every artifact produced by ``evaluate.py``
into a single ``.xlsx`` workbook.

For ``models/<version>/`` the script reads:
    evaluation.json
    residuals/<slug>.csv     (one per KPI)
    plots/<slug>_residuals.png  (optional; embedded if present)

…and writes:
    models/<version>/evaluation_report.xlsx

Sheets:
    01 Summary             headline metrics per KPI (train / test / CV)
    02 Residual stats      OOF residual statistics per KPI
    03 CV per-fold R²      fold-1..fold-5 R² + mean/std per KPI
    04 CV per-fold RMSE    fold-1..fold-5 RMSE + mean/std per KPI
    05 CV per-fold MAE     fold-1..fold-5 MAE + mean/std per KPI
    06 Prediction intervals  90% conformal + GPR-native (where applicable)
    07 Methodology         human-readable notes
    <slug>  ×20            per-KPI sheet:
                              · metadata + all metrics
                              · per-fold CV table
                              · residual stats
                              · prediction-interval rows
                              · embedded residual plot (if PNG exists)
                              · full residuals table (row, y_true, y_pred_oof, residual)

Run:
    python evaluate_to_excel.py [--version v1]
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[1]

# ───────────────────────── styling ─────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color="1F4E79", size=14)
SUBHEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
SUBHEAD_FONT = Font(bold=True, color="1F4E79", size=11)
KEY_FONT = Font(bold=True)
GOOD_FILL = PatternFill("solid", fgColor="E2EFDA")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
BAD_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin", color="C0C0C0")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def _safe(v: Any) -> Any:
    """Return v unchanged unless it's NaN/inf → return None for Excel."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            if math.isnan(v) or math.isinf(v):
                return None
        except (TypeError, ValueError):
            return None
    return v


def _fmt_num(v: Any, ndigits: int = 4) -> Any:
    v = _safe(v)
    if isinstance(v, (int, float)):
        return round(float(v), ndigits)
    return v


def _confidence_fill(r2: Optional[float]) -> Optional[PatternFill]:
    if r2 is None or not isinstance(r2, (int, float)) or math.isnan(r2):
        return None
    if r2 >= 0.85:
        return GOOD_FILL
    if r2 >= 0.65:
        return WARN_FILL
    return BAD_FILL


def _safe_sheet_name(name: str) -> str:
    bad = set('[]:*?/\\')
    cleaned = "".join("_" if c in bad else c for c in name)[:31]
    return cleaned or "sheet"


def _write_header_row(ws: Worksheet, row: int, headers: List[str], col_offset: int = 1) -> None:
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_offset + i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER


def _write_subhead(ws: Worksheet, row: int, label: str, span: int = 1, col_offset: int = 1) -> None:
    c = ws.cell(row=row, column=col_offset, value=label)
    c.fill = SUBHEAD_FILL
    c.font = SUBHEAD_FONT
    c.alignment = LEFT
    if span > 1:
        ws.merge_cells(
            start_row=row, start_column=col_offset,
            end_row=row, end_column=col_offset + span - 1,
        )


def _autosize(ws: Worksheet, max_width: int = 38) -> None:
    for col_cells in ws.columns:
        col_letter = None
        max_len = 0
        for cell in col_cells:
            if cell.column_letter is None:
                continue
            col_letter = cell.column_letter
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        if col_letter is not None:
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


# ───────────────────── builders for the 7 overview sheets ─────────────────────
def build_summary_sheet(wb: Workbook, report: Dict[str, Any]) -> None:
    ws = wb.create_sheet("01 Summary")
    ws.cell(row=1, column=1, value="NOLHC ML surrogate — headline metrics per KPI").font = TITLE_FONT
    ws.cell(
        row=2, column=1,
        value=(
            f"version={report['version']}  ·  n={report['n_total']} "
            f"(train={report['n_train']} / test={report['n_test']})  ·  "
            f"CV={report['n_splits_cv']}-fold seed={report['random_state']}  ·  "
            f"PI={int((1 - report['alpha_prediction_interval']) * 100)}% (split-conformal on OOF residuals)"
        ),
    ).font = Font(italic=True, color="595959")

    headers = [
        "#", "KPI slug", "Raw key", "Unit", "Registered model",
        "Train RMSE", "Train MAE", "Train R²",
        "Test RMSE", "Test MAE", "Test R²",
        "CV RMSE mean", "CV RMSE std",
        "CV MAE mean", "CV MAE std",
        "CV R² mean", "CV R² std",
        "Pooled OOF RMSE", "Pooled OOF MAE", "Pooled OOF R²",
        "90% PI half-width (conformal)",
        "GPR median σ̂ (OOF)", "GPR empirical coverage",
    ]
    _write_header_row(ws, 4, headers)
    ws.freeze_panes = "F5"

    r = 5
    for i, (slug, o) in enumerate(report["outputs"].items(), start=1):
        tr = o["train_in_sample"]
        te = o["test_holdout"] if isinstance(o["test_holdout"], dict) else {}
        cv = o["cv"]
        pi = o["prediction_interval_90pct"]
        g = o.get("gpr_native_interval_90pct") or {}

        row = [
            i, slug, o["raw_key"], o["unit"], o["registered_as"],
            _fmt_num(tr["rmse"]), _fmt_num(tr["mae"]), _fmt_num(tr["r2"]),
            _fmt_num(te.get("rmse")), _fmt_num(te.get("mae")), _fmt_num(te.get("r2")),
            _fmt_num(cv.get("rmse_mean")), _fmt_num(cv.get("rmse_std")),
            _fmt_num(cv.get("mae_mean")), _fmt_num(cv.get("mae_std")),
            _fmt_num(cv.get("r2_mean")), _fmt_num(cv.get("r2_std")),
            _fmt_num((cv.get("pooled_oof") or {}).get("rmse")),
            _fmt_num((cv.get("pooled_oof") or {}).get("mae")),
            _fmt_num((cv.get("pooled_oof") or {}).get("r2")),
            _fmt_num(pi.get("half_width")) if "half_width" in pi else None,
            _fmt_num(g.get("median_std")),
            _fmt_num(g.get("empirical_coverage_oof_90pct")),
        ]
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=_safe(v))
            c.border = BORDER
            if j in (8, 11, 16):  # train R², test R², CV R²
                fill = _confidence_fill(v if isinstance(v, (int, float)) else None)
                if fill is not None:
                    c.fill = fill
            if j > 5:
                c.alignment = RIGHT
        r += 1

    _autosize(ws)


def build_residual_stats_sheet(wb: Workbook, report: Dict[str, Any]) -> None:
    ws = wb.create_sheet("02 Residual stats")
    ws.cell(row=1, column=1, value="Out-of-fold residual statistics (y − ŷ_OOF)").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Same KPI unit; computed on the 5-fold OOF predictions.").font = Font(
        italic=True, color="595959"
    )
    headers = [
        "#", "KPI slug", "Raw key", "Unit", "Model",
        "mean", "std", "|resid| mean", "min", "max", "q05", "q50", "q95",
    ]
    _write_header_row(ws, 4, headers)
    ws.freeze_panes = "F5"

    r = 5
    for i, (slug, o) in enumerate(report["outputs"].items(), start=1):
        rs = o["residuals_oof"]
        if "error" in rs:
            vals = [None] * 8
        else:
            vals = [
                _fmt_num(rs["mean"]), _fmt_num(rs["std"]), _fmt_num(rs["abs_mean"]),
                _fmt_num(rs["min"]), _fmt_num(rs["max"]),
                _fmt_num(rs["q05"]), _fmt_num(rs["q50"]), _fmt_num(rs["q95"]),
            ]
        row = [i, slug, o["raw_key"], o["unit"], o["registered_as"], *vals]
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=_safe(v))
            c.border = BORDER
            if j > 5:
                c.alignment = RIGHT
        r += 1

    _autosize(ws)


def _build_per_fold_sheet(
    wb: Workbook, report: Dict[str, Any], metric_key: str, sheet_name: str, title: str
) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    n_folds = report["n_splits_cv"]
    headers = ["#", "KPI slug", "Raw key", "Model"]
    headers += [f"fold {k+1}" for k in range(n_folds)]
    headers += ["mean", "std"]
    _write_header_row(ws, 3, headers)
    ws.freeze_panes = "E4"

    r = 4
    for i, (slug, o) in enumerate(report["outputs"].items(), start=1):
        cv = o["cv"]
        per_fold = cv.get("per_fold", []) or []
        fold_vals: List[Any] = []
        for k in range(n_folds):
            if k < len(per_fold) and "error" not in per_fold[k]:
                fold_vals.append(_fmt_num(per_fold[k].get(metric_key)))
            else:
                fold_vals.append(None)
        mean = _fmt_num(cv.get(f"{metric_key}_mean"))
        std = _fmt_num(cv.get(f"{metric_key}_std"))
        row = [i, slug, o["raw_key"], o["registered_as"], *fold_vals, mean, std]
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=_safe(v))
            c.border = BORDER
            if j > 4:
                c.alignment = RIGHT
        r += 1

    _autosize(ws)


def build_prediction_intervals_sheet(wb: Workbook, report: Dict[str, Any]) -> None:
    ws = wb.create_sheet("06 Prediction intervals")
    ws.cell(
        row=1, column=1,
        value=(
            f"{int((1 - report['alpha_prediction_interval']) * 100)}% prediction intervals — "
            "split-conformal (model-agnostic) and GPR-native (where applicable)."
        ),
    ).font = TITLE_FONT
    ws.cell(
        row=2, column=1,
        value=(
            "Conformal half-width = quantile(|y − ŷ_OOF|, ⌈0.90·(n+1)⌉/n). "
            "GPR-native half-width = 1.6449·σ̂(x) with σ̂ taken from OOF folds."
        ),
    ).font = Font(italic=True, color="595959")

    headers = [
        "#", "KPI slug", "Raw key", "Unit", "Model",
        "Conformal half-width", "n calibration",
        "GPR median σ̂ (OOF)", "GPR mean σ̂ (OOF)",
        "GPR median half-width", "GPR mean half-width",
        "GPR empirical coverage",
    ]
    _write_header_row(ws, 4, headers)
    ws.freeze_panes = "F5"

    r = 5
    for i, (slug, o) in enumerate(report["outputs"].items(), start=1):
        pi = o["prediction_interval_90pct"]
        g = o.get("gpr_native_interval_90pct") or {}
        row = [
            i, slug, o["raw_key"], o["unit"], o["registered_as"],
            _fmt_num(pi.get("half_width")) if "half_width" in pi else None,
            pi.get("n_calibration"),
            _fmt_num(g.get("median_std")),
            _fmt_num(g.get("mean_std")),
            _fmt_num(g.get("median_half_width")),
            _fmt_num(g.get("mean_half_width")),
            _fmt_num(g.get("empirical_coverage_oof_90pct")),
        ]
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=_safe(v))
            c.border = BORDER
            if j > 5:
                c.alignment = RIGHT
        r += 1

    _autosize(ws)


def build_methodology_sheet(wb: Workbook, report: Dict[str, Any]) -> None:
    ws = wb.create_sheet("07 Methodology")
    ws.cell(row=1, column=1, value="Methodology notes").font = TITLE_FONT
    notes = [
        ("Source", "Generated from models/<version>/evaluation.json by evaluate_to_excel.py."),
        ("Training (in-sample)",
         "The registered model (per-KPI winner) refit on the full dataset and scored on the same "
         "data. Use only for comparison vs. CV/test — this is the apparent error."),
        ("Test (holdout)",
         f"Single 80/20 split with random_state={report['random_state']}. A fresh clone of the "
         "registered architecture is refit on the 80% training rows and scored on the held-out 20%."),
        ("Cross-validation",
         f"{report['n_splits_cv']}-fold KFold(shuffle=True, random_state={report['random_state']}) — "
         "same seed as training time. Reported as per-fold values plus mean ± std and a pooled "
         "OOF metric."),
        ("Residual analysis",
         "All residuals are y − ŷ_OOF from the 5-fold pass. Per-row residuals live on each KPI's "
         "sheet (and as residuals/<slug>.csv on disk)."),
        ("Conformal prediction interval",
         f"Marginal {int((1 - report['alpha_prediction_interval']) * 100)}% interval built by "
         "split-conformal calibration on OOF |residuals| with the standard (n+1)/n correction. "
         "Model-agnostic; valid in distribution under exchangeability."),
        ("GPR-native interval",
         "For KPIs whose registered model is gpr_rbf or gpr_matern we additionally report the "
         "Gaussian interval ±1.6449·σ̂(x). σ̂ is collected on OOF points (held-out folds), so the "
         "summary reflects predictive uncertainty at unseen inputs. We also report the empirical "
         "coverage at the 90% level — values close to 0.90 indicate the GP's posterior std is "
         "well-calibrated."),
        ("Cell colouring on '01 Summary'",
         "Train R² / Test R² / CV R² cells are colour-banded: green ≥0.85, amber ≥0.65, red <0.65."),
    ]
    for i, (k, v) in enumerate(notes, start=3):
        kc = ws.cell(row=i, column=1, value=k); kc.font = KEY_FONT; kc.alignment = LEFT
        vc = ws.cell(row=i, column=2, value=v); vc.alignment = LEFT
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110


# ─────────────────────── per-KPI sheet builder ───────────────────────
def build_kpi_sheet(
    wb: Workbook,
    slug: str,
    info: Dict[str, Any],
    report: Dict[str, Any],
    residuals_csv: Path,
    plot_png: Optional[Path],
) -> None:
    ws = wb.create_sheet(_safe_sheet_name(slug))

    ws.cell(row=1, column=1, value=f"KPI: {info['raw_key']}  ({slug})").font = TITLE_FONT
    ws.cell(
        row=2, column=1,
        value=f"Unit: {info['unit']}   ·   Registered model: {info['registered_as']}   ·   n={report['n_total']}",
    ).font = Font(italic=True, color="595959")

    # Block 1 — training / test / CV metrics
    _write_subhead(ws, 4, "Train / Test / CV metrics", span=6)
    _write_header_row(ws, 5, ["View", "RMSE", "MAE", "R²", "Notes", ""])

    tr = info["train_in_sample"]
    te = info["test_holdout"] if isinstance(info["test_holdout"], dict) else {}
    cv = info["cv"]
    rows = [
        (
            "Training (full data, in-sample)",
            _fmt_num(tr["rmse"]), _fmt_num(tr["mae"]), _fmt_num(tr["r2"]),
            "Apparent error — registered model refit on all rows.", "",
        ),
        (
            f"Test (holdout {report['n_train']}/{report['n_test']})",
            _fmt_num(te.get("rmse")), _fmt_num(te.get("mae")), _fmt_num(te.get("r2")),
            f"Fresh clone fit on 80% (random_state={report['random_state']}), scored on 20%.", "",
        ),
        (
            f"CV {report['n_splits_cv']}-fold (mean ± std)",
            f"{_fmt_num(cv.get('rmse_mean'))} ± {_fmt_num(cv.get('rmse_std'))}",
            f"{_fmt_num(cv.get('mae_mean'))} ± {_fmt_num(cv.get('mae_std'))}",
            f"{_fmt_num(cv.get('r2_mean'))} ± {_fmt_num(cv.get('r2_std'))}",
            f"KFold(shuffle=True, random_state={report['random_state']}).", "",
        ),
        (
            "CV pooled OOF",
            _fmt_num((cv.get("pooled_oof") or {}).get("rmse")),
            _fmt_num((cv.get("pooled_oof") or {}).get("mae")),
            _fmt_num((cv.get("pooled_oof") or {}).get("r2")),
            "Metrics on the stitched-back OOF prediction vector.", "",
        ),
    ]
    r = 6
    for row in rows:
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=_safe(v))
            c.border = BORDER
            if j == 1:
                c.alignment = LEFT
                c.font = KEY_FONT
            elif j in (2, 3, 4):
                c.alignment = RIGHT
            else:
                c.alignment = LEFT
        r += 1

    # Block 2 — per-fold CV
    r += 1
    _write_subhead(ws, r, "Per-fold cross-validation", span=4)
    r += 1
    _write_header_row(ws, r, ["fold", "RMSE", "MAE", "R²"])
    r += 1
    per_fold = cv.get("per_fold", []) or []
    for k in range(report["n_splits_cv"]):
        if k < len(per_fold) and "error" not in per_fold[k]:
            row = [k + 1, _fmt_num(per_fold[k].get("rmse")),
                   _fmt_num(per_fold[k].get("mae")), _fmt_num(per_fold[k].get("r2"))]
        else:
            row = [k + 1, None, None, None]
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=_safe(v))
            c.border = BORDER
            c.alignment = RIGHT if j > 1 else CENTER
        r += 1
    for label, key in (("mean", "_mean"), ("std", "_std")):
        row = [label,
               _fmt_num(cv.get(f"rmse{key}")),
               _fmt_num(cv.get(f"mae{key}")),
               _fmt_num(cv.get(f"r2{key}"))]
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=_safe(v))
            c.border = BORDER
            c.font = KEY_FONT
            c.alignment = RIGHT if j > 1 else CENTER
        r += 1

    # Block 3 — residual stats
    r += 1
    _write_subhead(ws, r, "Residual statistics (out-of-fold)", span=8)
    r += 1
    rs = info["residuals_oof"]
    _write_header_row(
        ws, r,
        ["mean", "std", "|resid| mean", "min", "max", "q05", "q50", "q95"],
    )
    r += 1
    if "error" in rs:
        vals = [None] * 8
    else:
        vals = [
            _fmt_num(rs["mean"]), _fmt_num(rs["std"]), _fmt_num(rs["abs_mean"]),
            _fmt_num(rs["min"]), _fmt_num(rs["max"]),
            _fmt_num(rs["q05"]), _fmt_num(rs["q50"]), _fmt_num(rs["q95"]),
        ]
    for j, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=j, value=_safe(v))
        c.border = BORDER
        c.alignment = RIGHT
    r += 2

    # Block 4 — prediction intervals
    _write_subhead(ws, r, "Prediction intervals (90%)", span=6)
    r += 1
    _write_header_row(
        ws, r,
        ["Method", "Half-width", "n calibration", "Median σ̂", "Mean σ̂", "Empirical coverage"],
    )
    r += 1
    pi = info["prediction_interval_90pct"]
    ws.cell(row=r, column=1, value="Split-conformal (OOF residuals)").border = BORDER
    ws.cell(row=r, column=2, value=_safe(_fmt_num(pi.get("half_width")))).border = BORDER
    ws.cell(row=r, column=3, value=pi.get("n_calibration")).border = BORDER
    for col in (4, 5, 6):
        ws.cell(row=r, column=col, value=None).border = BORDER
    r += 1
    g = info.get("gpr_native_interval_90pct") or {}
    if g and "error" not in g:
        ws.cell(row=r, column=1, value="GPR-native (Gaussian, OOF σ̂)").border = BORDER
        ws.cell(row=r, column=2, value=_safe(_fmt_num(g.get("median_half_width")))).border = BORDER
        ws.cell(row=r, column=3, value=None).border = BORDER
        ws.cell(row=r, column=4, value=_safe(_fmt_num(g.get("median_std")))).border = BORDER
        ws.cell(row=r, column=5, value=_safe(_fmt_num(g.get("mean_std")))).border = BORDER
        ws.cell(row=r, column=6, value=_safe(_fmt_num(g.get("empirical_coverage_oof_90pct")))).border = BORDER
        r += 1

    r += 1
    # Block 5 — embedded plot (if available)
    plot_anchor_row = r
    if plot_png is not None and plot_png.is_file():
        _write_subhead(ws, r, "Residual plot (pred-vs-actual · residual-vs-fitted)", span=8)
        r += 1
        try:
            img = XLImage(str(plot_png))
            img.width = 720
            img.height = 290
            ws.add_image(img, f"A{r}")
            r += 16
        except Exception as exc:  # noqa: BLE001
            ws.cell(row=r, column=1, value=f"(failed to embed image: {exc})")
            r += 1

    # Block 6 — full residuals data
    r += 1
    _write_subhead(ws, r, "OOF residuals (one row per simulation run)", span=4)
    r += 1
    if residuals_csv.is_file():
        df = pd.read_csv(residuals_csv)
        _write_header_row(ws, r, list(df.columns))
        r += 1
        for _, srow in df.iterrows():
            for j, col in enumerate(df.columns, start=1):
                c = ws.cell(row=r, column=j, value=_safe(_fmt_num(srow[col], 6)))
                c.border = BORDER
                if j > 1:
                    c.alignment = RIGHT
            r += 1
    else:
        ws.cell(row=r, column=1, value=f"(residual CSV not found at {residuals_csv})")

    # Column widths
    for col_idx, width in enumerate(
        [34, 14, 14, 14, 60, 14, 14, 14], start=1
    ):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A4"


# ───────────────────────────── main entry ─────────────────────────────
def to_excel(version: str = "v1") -> Path:
    model_dir = PROJECT_ROOT / "models" / version
    if not model_dir.is_dir():
        raise FileNotFoundError(f"Missing model directory: {model_dir}")

    eval_json = model_dir / "evaluation.json"
    if not eval_json.is_file():
        raise FileNotFoundError(
            f"Missing {eval_json}. Run `python evaluate.py --version {version}` first."
        )

    residuals_dir = model_dir / "residuals"
    plots_dir = model_dir / "plots"

    with open(eval_json, encoding="utf-8") as f:
        report: Dict[str, Any] = json.load(f)

    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    build_summary_sheet(wb, report)
    build_residual_stats_sheet(wb, report)
    _build_per_fold_sheet(wb, report, "r2", "03 CV per-fold R²", "Per-fold cross-validation — R²")
    _build_per_fold_sheet(wb, report, "rmse", "04 CV per-fold RMSE", "Per-fold cross-validation — RMSE")
    _build_per_fold_sheet(wb, report, "mae", "05 CV per-fold MAE", "Per-fold cross-validation — MAE")
    build_prediction_intervals_sheet(wb, report)
    build_methodology_sheet(wb, report)

    for slug, info in report["outputs"].items():
        csv_path = residuals_dir / f"{slug}.csv"
        png_path = plots_dir / f"{slug}_residuals.png"
        build_kpi_sheet(
            wb, slug, info, report,
            residuals_csv=csv_path,
            plot_png=png_path if png_path.is_file() else None,
        )

    out_path = model_dir / "evaluation_report.xlsx"
    wb.save(out_path)
    return out_path


def main(argv: Optional[List[str]] = None) -> None:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--version", default="v1", help="model version directory under models/")
    args = p.parse_args(argv)
    out = to_excel(args.version)
    print(f"[evaluate_to_excel] wrote {out}")


if __name__ == "__main__":
    main(sys.argv[1:])
